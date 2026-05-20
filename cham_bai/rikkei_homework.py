"""
Chấm BTVN qua API Rikkei Academy: lấy danh sách học sinh, bài tập, nhận xét (OpenRouter),
tuỳ chọn ghi comment lên portal, xuất Excel.
"""

from __future__ import annotations

import os
import re
import math
import unicodedata
import random
import tempfile
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from cham_bai.collector import CollectedBundle, format_bundle_for_prompt
from cham_bai.git_remote import fetch_repo_sources_bundle, normalize_github_repo_url
from cham_bai.google_sheets import _norm_name_for_match
from cham_bai.openrouter import ChatMessage, complete_chat


RIKKEI_BASE = "https://apiportal.rikkei.edu.vn"

# Portal thường kiểm tra từ trình duyệt; thiếu UA đôi khi bị đóng kết nối sớm.
_RIKKEI_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
)

_RETRYABLE_EXC = (
    httpx.RemoteProtocolError,
    httpx.LocalProtocolError,
    httpx.ConnectError,
    httpx.ReadTimeout,
    httpx.WriteTimeout,
    httpx.ConnectTimeout,
    httpx.PoolTimeout,
    BrokenPipeError,
    ConnectionResetError,
)

_SCORE_FROM_COMMENT_RE = re.compile(
    # Ví dụ: "[Điểm số - 90/100]" hoặc biến thể tương tự.
    # Chỉ cần bắt số đứng trước "/100" trong cặp ngoặc vuông.
    r"\[\s*[^]]*?(?P<score>\d+(?:\.\d+)?)\s*/\s*100\s*\]",
    re.IGNORECASE,
)

# Portal mới: "Kết quả: ✔ ĐẠT — 90/100" (không dùng ngoặc vuông).
_KET_QUA_SCORE_RE = re.compile(
    r"Kết\s*quả\s*:\s*[^\n]*?(?P<score>\d+(?:\.\d+)?)\s*/\s*100",
    re.IGNORECASE,
)
# Cùng dòng có ĐẠT và điểm (dự phòng khi thiếu tiền tố "Kết quả:").
_DAT_LINE_SCORE_RE = re.compile(
    r"(?:✔\s*)?(?:ĐẠT|đạt)\s*[—\-–]?\s*(?P<score>\d+(?:\.\d+)?)\s*/\s*100",
    re.IGNORECASE,
)

# Bài Mindmap: không chấm điểm; vẫn tính +1 bài hoàn thành; nhận xét cố định.
MINDMAP_HOMEWORK_PHRASE = "Hệ thống kiến thức Mindmap"
MINDMAP_PLACEHOLDER_COMMENT = "Chưa có nhận xét"


def _scores_from_comment(comment: str | None) -> list[float]:
    c = str(comment or "")
    scores: list[float] = []
    for rx in (_SCORE_FROM_COMMENT_RE, _KET_QUA_SCORE_RE, _DAT_LINE_SCORE_RE):
        for m in rx.finditer(c):
            try:
                scores.append(float(m.group("score")))
            except Exception:
                continue
    return scores


def _score_from_comment(comment: str | None) -> float | None:
    """
    Trả về điểm cao nhất tìm được trong nhận xét (một bài có thể ghép nhiều khối "Kết quả:").
    """
    s = _scores_from_comment(comment)
    return max(s) if s else None


_KET_QUA_LINE_START_RE = re.compile(r"(?i)^\s*Kết\s*quả\s*:")


def _ai_detected_in_comment(comment: str | None) -> bool:
    """
    Portal đôi khi có cảnh báo AI (tiếng Việt/Anh). Nếu có, cộng +10 vào điểm trước khi so ngưỡng đạt.
    """
    c = str(comment or "")
    if not c.strip():
        return False
    low = c.lower()
    if "ai detected" in low:
        return True
    if "nghi" in low and "ai" in low:
        # "Nghi dùng AI", "Nghi ngờ dùng AI: Có", ...
        return True
    if "⚠" in c:
        return True
    return False


def _ket_qua_dat_decision(comment: str | None) -> bool | None:
    """
    Định dạng portal: dòng bắt đầu bằng "Kết quả: ...".
    - True: có ít nhất một dòng ĐẠT (không phải CHƯA ĐẠT).
    - False: có CHƯA ĐẠT và không có dòng ĐẠT thuần nào.
    - None: không có dòng Kết quả:, hoặc có nhưng không suy ra được → dùng điểm.
    """
    c = str(comment or "")
    lines = [ln for ln in c.splitlines() if _KET_QUA_LINE_START_RE.match(ln)]
    if not lines:
        return None
    found_pass = False
    found_chua = False
    for line in lines:
        if re.search(r"(?i)chưa\s*đạt", line):
            found_chua = True
            continue
        if re.search(r"(?i)ĐẠT|\bđạt\b", line):
            found_pass = True
    if found_pass:
        return True
    if found_chua:
        return False
    return None


def _exercise_achieved_for_session(comment: str | None, score_threshold: float) -> bool:
    """Một bài tính là đạt khi portal ghi rõ ĐẠT, hoặc (fallback) điểm > ngưỡng."""
    d = _ket_qua_dat_decision(comment)
    if d is True:
        return True
    if d is False:
        return False
    sc = _score_from_comment(comment)
    if sc is None:
        return False
    if _ai_detected_in_comment(comment):
        sc = min(100.0, sc + 10.0)
    return sc > score_threshold


def _extract_student_session_status(student: dict[str, Any], session_id: int | str | None = None) -> str:
    """
    Rikkei có nhiều biến thể response.
    `sessionStudent` thường là mảng nhiều session, cần pick đúng theo session_id (nếu có).
    """
    sid = str(session_id).strip() if session_id is not None else ""
    ss = student.get("sessionStudent")
    if isinstance(ss, dict):
        st = ss.get("status")
        if isinstance(st, str) and st.strip():
            return st.strip()
    if isinstance(ss, list) and ss:
        if sid:
            for it in ss:
                if not isinstance(it, dict):
                    continue
                it_sid = it.get("sessionId") or it.get("session_id")
                if it_sid is None and isinstance(it.get("session"), dict):
                    it_sid = it["session"].get("id")
                if it_sid is not None and str(it_sid).strip() == sid:
                    st = it.get("status")
                    if isinstance(st, str) and st.strip():
                        return st.strip()
            # Nếu có session_id nhưng không match được item nào (API không trả sessionId),
            # thì KHÔNG fallback bừa sang phần tử đầu (có thể là session khác).
        else:
            # Không có session_id -> best-effort: lấy status đầu tiên có giá trị
            for it in ss:
                if not isinstance(it, dict):
                    continue
                st = it.get("status")
                if isinstance(st, str) and st.strip():
                    return st.strip()

    # Fallback: các trường top-level
    for k in ("status", "homeworkStatus", "homework_status", "sessionStatus", "session_status"):
        v = student.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()

    return ""


def _norm_status_text(s: str) -> str:
    """
    Chuẩn hóa chuỗi để so sánh trạng thái không bị lệch dấu tiếng Việt.
    """
    t = str(s or "")
    # Strip diacritics
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    return t.strip().upper()


def _extract_strengths_weaknesses(comment: str) -> str:
    """
    Lấy gọn 2 mục: "Điểm mạnh" và "Điểm yếu" (nếu có) từ comment rubric.
    Trả về chuỗi gọn để ghi vào sheet.
    """
    raw = str(comment or "").strip()
    if not raw:
        return ""

    # Portal có thể trả HTML (div/p/span/hr...) và đôi khi bị cụt giữa chừng.
    # Cần strip robust để luôn lấy plain text, bỏ block "Kết quả/điểm".
    if "<" in raw and ">" in raw:
        try:
            from html import unescape as _html_unescape

            s = raw
            # Keep boundaries even if missing closing tags
            s = re.sub(r"(?is)<\s*br\s*/?\s*>", "\n", s)
            s = re.sub(r"(?is)<\s*/\s*p\s*>", "\n\n", s)
            s = re.sub(r"(?is)<\s*p\b[^>]*>", "\n\n", s)
            s = re.sub(r"(?is)<\s*hr\b[^>]*>", "\n\n", s)
            # Strip all tags (best-effort even for truncated HTML)
            s = re.sub(r"(?is)<[^>]*>", " ", s)
            s = _html_unescape(s)

            # Normalize whitespace
            s = s.replace("\u00a0", " ")
            s = re.sub(r"[ \t\r\f\v]+", " ", s)
            s = re.sub(r"\n[ \t]+", "\n", s)
            s = re.sub(r"\n{3,}", "\n\n", s).strip()

            # Remove "Kết quả" block + score patterns
            lines = [ln.strip() for ln in s.splitlines()]
            kept: list[str] = []
            for ln in lines:
                if not ln:
                    kept.append("")
                    continue
                low_ln = ln.lower()
                if "kết quả" in low_ln:
                    continue
                if "đạt" in low_ln and ("✔" in ln or "✘" in ln):
                    continue
                if re.search(r"\b\d+\s*/\s*100\b", ln):
                    continue
                kept.append(ln)
            s2 = "\n".join(kept)
            s2 = re.sub(r"\n{3,}", "\n\n", s2).strip()

            # Prefer longest paragraph; fallback to whole
            paras = [p.strip() for p in re.split(r"\n\s*\n", s2) if p.strip()]
            raw = max(paras, key=len) if paras else (s2 or raw)
        except Exception:
            # worst case: aggressively strip tags without unescape
            s = re.sub(r"(?is)<[^>]*>", " ", raw)
            s = re.sub(r"[ \t\r\f\v]+", " ", s).strip()
            if s:
                raw = s

    # Loại bỏ các đoạn cảnh báo AI / tiếng Anh dài (portal đôi khi trả kèm "AI detected ...").
    # Mục tiêu sheet: chỉ giữ nhận xét tiếng Việt ngắn gọn.
    def _has_vn_marks(t: str) -> bool:
        return bool(
            re.search(
                r"[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ"
                r"ÀÁẠẢÃÂẦẤẬẨẪĂẰẮẶẲẴÈÉẸẺẼÊỀẾỆỂỄÌÍỊỈĨÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠÙÚỤỦŨƯỪỨỰỬỮỲÝỴỶỸĐ]",
                t,
            )
        )

    raw_lines = [ln.strip() for ln in raw.splitlines()]
    kept2: list[str] = []
    for ln in raw_lines:
        if not ln:
            kept2.append("")
            continue
        low_ln = ln.lower()
        if low_ln.startswith("⚠") or "ai detected" in low_ln:
            continue
        if ("nghi" in low_ln and "ai" in low_ln) or ("ngờ" in low_ln and "ai" in low_ln):
            # ví dụ: "Nghi dùng AI: ..." / "Nghi ngờ dùng AI: ..."
            continue
        # Nếu là tiếng Anh dài (không có dấu tiếng Việt) thì bỏ.
        if not _has_vn_marks(ln) and len(ln) >= 60:
            continue
        kept2.append(ln)
    raw2 = "\n".join(kept2)
    raw2 = re.sub(r"\n{3,}", "\n\n", raw2).strip()
    if raw2:
        raw = raw2

    low = raw.lower()
    # Find anchors (Vietnamese with/without accents already handled upstream in sheet matching)
    idx_s = low.find("điểm mạnh")
    idx_w = low.find("điểm yếu")
    if idx_s < 0 and idx_w < 0:
        # fallback: keep first 400 chars
        return (raw[:400].rstrip() + "…") if len(raw) > 400 else raw

    # Slice from "Điểm mạnh" to end or to before the next big section.
    start = idx_s if idx_s >= 0 else idx_w
    tail = raw[start:]
    # Stop at "Gợi ý" / "Suggestion" markers if present.
    cut_markers = ["\nGợi ý", "\nGỌI Ý", "\nGoi y", "\nHướng dẫn", "\nTóm tắt", "\nYêu cầu"]
    end = len(tail)
    low_tail = tail.lower()
    for mk in cut_markers:
        j = low_tail.find(mk.lower())
        if j > 0:
            end = min(end, j)
    tail = tail[:end].strip()

    # If both sections exist, keep both.
    if idx_s >= 0 and idx_w >= 0 and idx_w > idx_s:
        # already included, just ensure it contains both headings
        pass
    elif idx_s < 0 and idx_w >= 0:
        # only weaknesses
        pass
    elif idx_s >= 0 and idx_w < 0:
        # only strengths
        pass

    # Normalize whitespace a bit
    tail = re.sub(r"\n{3,}", "\n\n", tail).strip()
    return tail


def _norm_text_simple(s: str) -> str:
    t = str(s or "")
    t = unicodedata.normalize("NFD", t)
    t = "".join(ch for ch in t if unicodedata.category(ch) != "Mn")
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


_BTVN_TIER_END_RE = re.compile(
    r"\s*(?P<tier>yếu|yeu|tb|khá|kha|giỏi|gioi)\s*$",
    re.IGNORECASE,
)
_BTVN_DATE_CHUNK_RE = re.compile(
    r"(?:[\s_\-–—]+)?(?:\(?\s*)?\d{1,2}\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{2,4}(?:\s*\)?)?",
    re.IGNORECASE,
)
_BTVN_GROUP_LINE_RE = re.compile(
    r"^(?:cntt\s*\d*|nh[oó]m\s*(?:cá\s*biệt|\d+)|group\s*\d*)\s*$",
    re.IGNORECASE,
)


def _canonical_btvn_tier(raw: str) -> str | None:
    t = _norm_text_simple(raw)
    if t in ("yeu", "yếu"):
        return "Yếu"
    if t == "tb":
        return "TB"
    if t in ("kha", "khá"):
        return "Khá"
    if t in ("gioi", "giỏi"):
        return "Giỏi"
    return None


def parse_btvn_student_tier_text(text: str) -> dict[str, str]:
    """
    Parse danh sách phân loại (mỗi dòng: họ tên + ngày sinh tuỳ chọn + Yếu/TB/Khá/Giỏi).
    Trả về map tên đã chuẩn hoá -> tier chuẩn.
    """
    out: dict[str, str] = {}
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if _BTVN_GROUP_LINE_RE.match(line):
            continue
        m = _BTVN_TIER_END_RE.search(line)
        if not m:
            continue
        tier = _canonical_btvn_tier(m.group("tier"))
        if not tier:
            continue
        name_part = line[: m.start()].strip()
        name_part = _BTVN_DATE_CHUNK_RE.sub(" ", name_part)
        name_part = re.sub(r"[\s_\-–—]+$", "", name_part).strip()
        name_part = re.sub(r"\s+", " ", name_part).strip()
        if not name_part:
            continue
        key = _norm_name_for_match(name_part)
        if key:
            out[key] = tier
    return out


def lookup_btvn_student_tier(full_name: str, tiers: dict[str, str]) -> str | None:
    key = _norm_name_for_match(full_name)
    if not key or not tiers:
        return None
    if key in tiers:
        return tiers[key]
    best: tuple[int, str] | None = None
    for k, tier in tiers.items():
        if not k:
            continue
        if k == key:
            return tier
        if len(k) >= 6 and (k in key or key in k):
            score = min(len(k), len(key))
            if best is None or score > best[0]:
                best = (score, tier)
    return best[1] if best else None


def btvn_tier_required_indices(tier: str, total: int) -> list[int]:
    """
    Chỉ số bài (0-based, theo thứ tự homework session) bắt buộc phải đạt:
    - Yếu: 2 bài đầu
    - TB: 3 bài đầu
    - Khá: 3 bài (bỏ qua bài đầu)
    - Giỏi: 3 bài cuối
    """
    if total <= 0:
        return []
    t = _norm_text_simple(tier)
    if t in ("yeu", "yếu"):
        return list(range(min(2, total)))
    if t == "tb":
        return list(range(min(3, total)))
    if t in ("kha", "khá"):
        start = 1 if total > 1 else 0
        return list(range(start, min(start + 3, total)))
    if t in ("gioi", "giỏi"):
        if total >= 3:
            return list(range(total - 3, total))
        return list(range(total))
    return list(range(min(3, total)))


def fetch_homework_session_order(token: str, session_id: int | str) -> list[int]:
    """Thứ tự homework id trong session (dùng sắp xếp exercise)."""
    sid = str(session_id).strip()
    ids: list[int] = []
    url = f"{RIKKEI_BASE}/homework/session/{sid}"
    try:
        r = _rikkei_request("GET", url, token)
        r.raise_for_status()
        data = r.json()
        items: list[Any] = []
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict):
            for k in ("homework", "homeworks", "data", "items", "content"):
                v = data.get(k)
                if isinstance(v, list):
                    items = v
                    break
            if not items:
                items = _unwrap_array(data)
        for item in items:
            if not isinstance(item, dict):
                continue
            hid = _as_int(item.get("id"))
            if hid is not None:
                ids.append(hid)
    except Exception:
        pass
    if ids:
        return ids
    try:
        r2 = _rikkei_request("GET", f"{RIKKEI_BASE}/sessions/{sid}", token)
        r2.raise_for_status()
        payload = r2.json()
        if isinstance(payload, dict):
            hw = payload.get("homework")
            if isinstance(hw, list):
                for item in hw:
                    if isinstance(item, dict):
                        hid = _as_int(item.get("id"))
                        if hid is not None:
                            ids.append(hid)
    except Exception:
        pass
    return ids


def sort_exercises_by_session_homework(
    exercises: list[dict[str, Any]],
    homework_ids_ordered: list[int],
) -> list[dict[str, Any]]:
    if not homework_ids_ordered:
        return list(exercises)
    hw_to_idx = {hid: i for i, hid in enumerate(homework_ids_ordered)}

    def _key(ex: dict[str, Any]) -> tuple[int, int, int]:
        hid = _homework_id(ex)
        if hid is None:
            return (1, 9999, _as_int(ex.get("id")) or 0)
        return (0, hw_to_idx.get(hid, 9999), hid)

    return sorted(exercises, key=_key)


def fetch_homework_session_total(token: str, session_id: int | str) -> int:
    ordered = fetch_homework_session_order(token, session_id)
    if ordered:
        return len(ordered)
    sid = str(session_id).strip()
    url = f"{RIKKEI_BASE}/homework/session/{sid}"
    r = _rikkei_request("GET", url, token)
    r.raise_for_status()
    data = r.json()
    if isinstance(data, dict):
        for k in ("total", "count", "totalExercises", "total_exercises"):
            v = data.get(k)
            if isinstance(v, int):
                return v
            if isinstance(v, str) and v.strip().isdigit():
                return int(v.strip())
    # fallback
    return 0


def _exercise_achieved_with_mindmap_side_effects(
    token: str,
    ex: dict[str, Any],
    *,
    score_threshold: float,
) -> bool:
    if is_mindmap_homework_exercise(ex):
        eid = _as_int(ex.get("id"))
        link = ex.get("link_git") or ex.get("linkGit") or ""
        if isinstance(link, str):
            link = link.strip()
        else:
            link = ""
        if eid is not None and link:
            try:
                put_exercise_comment(
                    token,
                    eid,
                    comment=MINDMAP_PLACEHOLDER_COMMENT,
                    link_git=link,
                    homework_id=_homework_id(ex),
                    full_body=ex,
                )
            except Exception:
                pass
        return True
    cmt = ex.get("comment")
    if isinstance(cmt, str) and cmt.strip():
        return _exercise_achieved_for_session(cmt.strip(), score_threshold)
    return False


def session_student_update(
    token: str,
    *,
    student_id: int | str,
    session_id: int | str,
    status: str,
    completed_exercises: int,
) -> None:
    sid = int(str(session_id).strip())
    stid = int(str(student_id).strip())
    body = {
        "studentId": stid,
        "sessionId": sid,
        "status": status,
        "completedExercises": int(completed_exercises),
    }
    r = _rikkei_request("POST", f"{RIKKEI_BASE}/session-student", token, json=body)
    if r.status_code >= 400:
        try:
            detail = r.json()
        except Exception:
            detail = r.text
        raise RuntimeError(f"Rikkei session-student error: {r.status_code} {detail}")


def resolve_btvn_pass_required_count(
    total: int,
    *,
    min_completed: int | None = None,
    ratio_ok: float = 0.5,
) -> int:
    """Số bài đạt tối thiểu để gán HOÀN THÀNH (ưu tiên min_completed nếu có)."""
    if total <= 0:
        return 1
    if min_completed is not None and int(min_completed) > 0:
        return min(int(min_completed), total)
    n = int(math.ceil(total * float(ratio_ok) - 1e-12))
    return max(1, min(n, total))


def mark_btvn_session_status_from_exercise_scores(
    token: str,
    *,
    class_id: int | str,
    course_id: int | str,
    session_id: int | str,
    student_ids: list[int] | None = None,
    score_threshold: float = 50.0,
    ratio_ok: float = 0.5,
    min_completed: int | None = None,
    student_tiers: dict[str, str] | None = None,
    waiting_status: str = "ĐANG CHỜ KIỂM TRA",
) -> dict[str, Any]:
    """
    Chốt trạng thái session dựa vào comment của từng exercise.
    - Điểm > score_threshold hoặc dòng ĐẠT => đạt từng bài
    - Chế độ mặc định: đạt >= pass_required (min_completed hoặc ceil(ratio_ok * total))
    - Chế độ phân loại (student_tiers): Yếu 2 đầu, TB 3 đầu, Khá 3 (bỏ bài 1), Giỏi 3 cuối — tất cả slot bắt buộc phải đạt
    - Chỉ cập nhật nếu trạng thái hiện tại = waiting_status
    """
    sid = str(session_id).strip()
    hw_order = fetch_homework_session_order(token, sid)
    total = len(hw_order) if hw_order else fetch_homework_session_total(token, sid)
    if total <= 0:
        return {"ok": False, "reason": "Không đọc được total bài trong homework/session.", "total": total}

    use_tier_mode = bool(student_tiers)
    ratio_ok_count = resolve_btvn_pass_required_count(
        total, min_completed=min_completed, ratio_ok=ratio_ok
    )

    students = fetch_students(token, class_id, sid)
    if student_ids:
        want = set(int(x) for x in student_ids if x is not None)
        filtered = []
        for st in students:
            try:
                st_id_int = int(st.get("id"))
            except Exception:
                continue
            if st_id_int in want:
                filtered.append(st)
        students = filtered

    ok = 0
    fail = 0
    updated: list[dict[str, Any]] = []
    ignored: list[dict[str, Any]] = []

    for st in students:
        try:
            st_id = int(st.get("id"))
        except Exception:
            continue

        cur_status = _extract_student_session_status(st, sid)
        # Portal: null status => hiểu là "ĐANG CHỜ KIỂM TRA"
        if not str(cur_status or "").strip():
            cur_status = waiting_status

        achieved = 0
        tier_label: str | None = None
        required_slots: list[int] = []
        passed_slots = 0
        tier_missing = False
        comments_pool: list[str] = []
        try:
            exercises = fetch_all_exercises_for_student(
                token,
                class_id=str(class_id),
                course_id=str(course_id),
                session_id=str(sid),
                student_id=st_id,
                page_limit=100,
            )
            exercises = sort_exercises_by_session_homework(exercises, hw_order)
            slot_achieved: list[bool] = []
            for ex in exercises:
                if not isinstance(ex, dict):
                    continue
                cmt = ex.get("comment")
                if isinstance(cmt, str) and cmt.strip():
                    comments_pool.append(cmt.strip())
                ok_slot = _exercise_achieved_with_mindmap_side_effects(
                    token, ex, score_threshold=score_threshold
                )
                slot_achieved.append(ok_slot)

            achieved = sum(1 for x in slot_achieved if x)

            if use_tier_mode and student_tiers:
                st_name = str(st.get("fullName") or st.get("full_name") or "").strip()
                tier_label = lookup_btvn_student_tier(st_name, student_tiers)
                if not tier_label:
                    tier_missing = True
                    new_status = "CHƯA HOÀN THÀNH"
                    required_slots = []
                    passed_slots = 0
                else:
                    required_slots = btvn_tier_required_indices(tier_label, total)
                    passed_slots = 0
                    for idx in required_slots:
                        if idx < len(slot_achieved) and slot_achieved[idx]:
                            passed_slots += 1
                    new_status = (
                        "HOÀN THÀNH"
                        if required_slots and passed_slots >= len(required_slots)
                        else "CHƯA HOÀN THÀNH"
                    )
                    achieved = passed_slots
            else:
                new_status = "HOÀN THÀNH" if achieved >= ratio_ok_count else "CHƯA HOÀN THÀNH"
        except Exception:
            fail += 1
            updated.append({"studentId": st_id, "ok": False, "error": "Lỗi đọc exercise để tính điểm"})
            continue
        rand_comment_raw = random.choice(comments_pool) if comments_pool else ""
        rand_comment = _extract_strengths_weaknesses(rand_comment_raw)

        cur_norm = _norm_status_text(cur_status)
        wait_norm = _norm_status_text(waiting_status)
        # Chỉ POST portal khi đang ở trạng thái chờ kiểm tra; nhưng vẫn trả dữ liệu để fill sheet.
        should_post = not (cur_norm and (cur_norm != wait_norm) and (wait_norm not in cur_norm))
        extra_row = {
            "tier": tier_label,
            "requiredSlots": required_slots,
            "passedSlots": passed_slots,
            "tierMissing": tier_missing,
        }
        if not should_post:
            ignored.append(
                {
                    "studentId": st_id,
                    "status": cur_status,
                    "ignoredBecause": "not_waiting",
                    "newStatus": new_status,
                    "completedExercises": achieved,
                    "achieved": achieved,
                    "randomComment": rand_comment,
                    "total": total,
                    **extra_row,
                }
            )
            continue
        try:
            session_student_update(
                token,
                student_id=st_id,
                session_id=sid,
                status=new_status,
                completed_exercises=achieved,
            )
            ok += 1
            updated.append(
                {
                    "studentId": st_id,
                    "ok": True,
                    "newStatus": new_status,
                    "completedExercises": achieved,
                    "achieved": achieved,
                    "randomComment": rand_comment,
                    "total": total,
                    **extra_row,
                }
            )
        except Exception as e:
            fail += 1
            updated.append(
                {
                    "studentId": st_id,
                    "ok": False,
                    "newStatus": new_status,
                    "completedExercises": achieved,
                    "achieved": achieved,
                    "randomComment": rand_comment,
                    "error": str(e)[:200],
                    **extra_row,
                }
            )

    return {
        "ok": True,
        "total": total,
        "ok_count": ok,
        "fail_count": fail,
        "updated": updated[:50],
        "ignored": ignored[:50],
        "ratio_ok_count": ratio_ok_count,
        "min_completed_required": ratio_ok_count,
        "score_threshold": score_threshold,
        "ratio_ok": ratio_ok,
        "min_completed_override": min_completed,
        "tier_mode": use_tier_mode,
        "tier_students_parsed": len(student_tiers or {}),
    }


def _rikkei_client() -> httpx.Client:
    return httpx.Client(
        timeout=httpx.Timeout(connect=45.0, read=180.0, write=45.0, pool=60.0),
        limits=httpx.Limits(max_keepalive_connections=0, max_connections=10),
        http2=False,
        follow_redirects=True,
        verify=True,
    )


def _rikkei_headers(token: str, *, json_body: bool = False) -> dict[str, str]:
    t = (token or "").strip()
    h: dict[str, str] = {
        "Authorization": f"Bearer {t}",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": _RIKKEI_UA,
    }
    if json_body:
        h["Content-Type"] = "application/json"
    return h


def _rikkei_request(
    method: str,
    url: str,
    token: str,
    *,
    params: dict[str, Any] | None = None,
    json: dict[str, Any] | None = None,
    max_attempts: int = 4,
) -> httpx.Response:
    last: BaseException | None = None
    for attempt in range(1, max_attempts + 1):
        try:
            with _rikkei_client() as client:
                return client.request(
                    method,
                    url,
                    headers=_rikkei_headers(token, json_body=json is not None),
                    params=params,
                    json=json,
                )
        except _RETRYABLE_EXC as e:
            last = e
            if attempt < max_attempts:
                time.sleep(min(2 ** (attempt - 1), 8))
    assert last is not None
    tip = (
        " Gợi ý: SSH vào máy chủ, chạy `curl -sI https://apiportal.rikkei.edu.vn/` "
        "(egress HTTPS); kiểm tra firewall; token Bearer còn hạn (thử lại trên trình duyệt DevTools)."
    )
    raise RuntimeError(
        f"Gọi API Rikkei không ổn định sau {max_attempts} lần thử: {last!s}.{tip}"
    ) from last


def _unwrap_array(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for k in ("data", "items", "content", "records", "result"):
            v = payload.get(k)
            if isinstance(v, list):
                return v
    return []


def fetch_students(token: str, class_id: int | str, session_id: int | str) -> list[dict[str, Any]]:
    cid, sid = str(class_id).strip(), str(session_id).strip()
    url = f"{RIKKEI_BASE}/students/homeworkProcess/class/{cid}/session/{sid}"
    r = _rikkei_request("GET", url, token)
    r.raise_for_status()
    return [x for x in _unwrap_array(r.json()) if isinstance(x, dict)]


def fetch_exercises_page(
    token: str,
    *,
    class_id: str,
    course_id: str,
    session_id: str,
    student_id: int | str,
    page: int,
    limit: int,
) -> list[dict[str, Any]]:
    params: dict[str, Any] = {
        "classId": class_id or "",
        "courseId": course_id or "",
        "sessionId": str(session_id),
        "studentName": "",
        "page": page,
        "limit": limit,
        "studentId": str(student_id),
    }
    r = _rikkei_request("GET", f"{RIKKEI_BASE}/exercise", token, params=params)
    r.raise_for_status()
    return [x for x in _unwrap_array(r.json()) if isinstance(x, dict)]


def fetch_all_exercises_for_student(
    token: str,
    *,
    class_id: str,
    course_id: str,
    session_id: str,
    student_id: int | str,
    page_limit: int = 100,
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    page = 1
    while True:
        chunk = fetch_exercises_page(
            token,
            class_id=class_id,
            course_id=course_id,
            session_id=str(session_id),
            student_id=student_id,
            page=page,
            limit=page_limit,
        )
        if not chunk:
            break
        out.extend(chunk)
        if len(chunk) < page_limit:
            break
        page += 1
    return out


def _repo_group_key(link_git: str | None) -> str | None:
    if not (link_git or "").strip():
        return None
    norm = normalize_github_repo_url((link_git or "").strip())
    if not norm:
        return None
    return norm.rstrip("/").lower()


def _homework_id(ex: dict[str, Any]) -> int | None:
    hw = ex.get("homework")
    if isinstance(hw, dict):
        hid = hw.get("id")
        if isinstance(hid, int):
            return hid
        if isinstance(hid, str) and hid.isdigit():
            return int(hid)
    return None


def _homework_title(ex: dict[str, Any]) -> str:
    t = ex.get("homework_title")
    if isinstance(t, str) and t.strip():
        return t.strip()
    hw = ex.get("homework")
    if isinstance(hw, dict):
        t = hw.get("title")
        if isinstance(t, str):
            return t.strip()
    return ""


def _exercise_label_blob(ex: dict[str, Any]) -> str:
    """Gộp các trường tiêu đề/mô tả để nhận diện bài đặc biệt (Mindmap)."""
    parts: list[str] = []
    for key in ("homework_title", "title", "name", "description"):
        v = ex.get(key)
        if isinstance(v, str) and v.strip():
            parts.append(v.strip())
    hw = ex.get("homework")
    if isinstance(hw, dict):
        for key in ("title", "name", "description"):
            v = hw.get(key)
            if isinstance(v, str) and v.strip():
                parts.append(v.strip())
    title = _homework_title(ex)
    if title:
        parts.append(title)
    return "\n".join(parts)


def is_mindmap_homework_exercise(ex: dict[str, Any]) -> bool:
    blob = _exercise_label_blob(ex)
    if not blob:
        return False
    return MINDMAP_HOMEWORK_PHRASE.casefold() in blob.casefold()


def _as_int(v: Any) -> int | None:
    if isinstance(v, int):
        return v
    if isinstance(v, str) and v.strip().isdigit():
        return int(v.strip())
    return None


def mark_shared_repo_violations(rows: list[dict[str, Any]]) -> None:
    """
    Cùng một repo GitHub gắn cho nhiều homework khác nhau → không coi là link riêng từng bài.
    """
    by_student: dict[int, dict[str, list[dict[str, Any]]]] = {}
    for r in rows:
        sid = r.get("student_id")
        if not isinstance(sid, int):
            continue
        key = r.get("_repo_key")
        if key is None:
            continue
        by_student.setdefault(sid, {}).setdefault(key, []).append(r)

    for _sid, buckets in by_student.items():
        for _key, group in buckets.items():
            hw_ids = {g.get("homework_id") for g in group if g.get("homework_id") is not None}
            if len(hw_ids) <= 1:
                continue
            for g in group:
                g["link_valid"] = False
                g["link_reason"] = (
                    "Một repo GitHub dùng chung cho nhiều bài tập khác nhau — không tính theo quy tắc link riêng từng bài."
                )


_BTVN_COMMENT_SYSTEM = """Bạn là giảng viên CNTT ở Việt Nam. Viết nhận xét BTVN cực ngắn (1–3 câu), tự nhiên như người thật, không markdown, không gạch đầu dòng.
Ưu tiên nhận xét dựa trên mã nguồn trong repo (nếu có), tập trung: đúng yêu cầu bài (theo tiêu đề), cấu trúc repo, chất lượng SQL/code, README, cách tổ chức file, tính rõ ràng.
Không chấm điểm, không đưa thang điểm. Giọng điềm tĩnh, giống người chấm thật."""


def _bundle_brief(bundle: CollectedBundle | None) -> str:
    if not bundle or not bundle.files:
        return "(Không đọc được mã nguồn từ repo.)"
    return format_bundle_for_prompt(bundle)


def generate_btvn_comment(
    *,
    homework_title: str,
    link_git: str,
    model: str,
    submission_bundle: CollectedBundle | None,
) -> str:
    user = (
        f"Tiêu đề bài: {homework_title or '(không rõ)'}\n"
        f"Link nộp: {link_git}\n\n"
        "Mã nguồn trong repo (có thể bị cắt bớt):\n"
        f"{_bundle_brief(submission_bundle)}\n\n"
        "Hãy viết nhận xét ngắn."
    )
    text, _ = complete_chat(
        [
            ChatMessage(role="system", content=_BTVN_COMMENT_SYSTEM),
            ChatMessage(role="user", content=user),
        ],
        model=model,
        temperature=0.45,
        max_tokens=450,
        timeout_s=120.0,
    )
    one_line = re.sub(r"\s+", " ", text).strip()
    if len(one_line) > 800:
        one_line = one_line[:797] + "..."
    return one_line


def get_exercise_detail(token: str, exercise_id: int | str) -> dict[str, Any] | None:
    r = _rikkei_request("GET", f"{RIKKEI_BASE}/exercise/{exercise_id}", token)
    if r.status_code == 404:
        return None
    r.raise_for_status()
    data = r.json()
    return data if isinstance(data, dict) else None


def put_exercise_comment(
    token: str,
    exercise_id: int | str,
    *,
    comment: str,
    link_git: str,
    homework_id: int | None,
    course_id: int | None = None,
    full_body: dict[str, Any] | None = None,
) -> tuple[bool, str]:
    cid: Any = course_id
    hw: Any = homework_id
    lg = link_git
    if isinstance(full_body, dict):
        cid = full_body.get("courseId", cid)
        if hw is None:
            hw = full_body.get("homeworkId")
        lg = (full_body.get("linkGit") or full_body.get("link_git") or lg) or lg
    body: dict[str, Any] = {
        "courseId": cid,
        "homeworkId": hw,
        "linkGit": lg,
        "comment": comment,
    }
    r = _rikkei_request(
        "PUT",
        f"{RIKKEI_BASE}/exercise/{exercise_id}",
        token,
        json=body,
    )
    if r.status_code >= 400:
        try:
            detail = r.json()
        except Exception:
            detail = r.text
        return False, str(detail)[:2000]
    return True, ""


@dataclass
class BtvnJobParams:
    rikkei_token: str
    class_id: str
    session_id: str
    course_id: str
    text_model: str
    push_to_portal: bool
    max_students: int | None
    openrouter_delay_s: float = 0.35


def _sanitize_filename_part(s: str) -> str:
    return re.sub(r"[^\w\-]+", "_", s, flags=re.UNICODE).strip("_") or "x"


def write_btvn_excel(rows: list[dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "BTVN"
    headers = [
        "Mã SV",
        "Họ tên",
        "studentId",
        "exerciseId",
        "homeworkId",
        "Tiêu đề bài",
        "link_git",
        "Link hợp lệ",
        "Lý do",
        "Lỗi đọc repo",
        "Nhận xét (AI)",
        "Lỗi AI",
        "Đã push portal",
        "Lỗi push",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for r in rows:
        ws.append(
            [
                r.get("student_code", ""),
                r.get("full_name", ""),
                r.get("student_id", ""),
                r.get("exercise_id", ""),
                r.get("homework_id", ""),
                r.get("homework_title", ""),
                r.get("link_git", ""),
                "Có" if r.get("link_valid") else "Không",
                r.get("link_reason", ""),
                r.get("repo_error", ""),
                r.get("ai_comment", ""),
                r.get("ai_error", ""),
                "Có" if r.get("pushed") else "Không",
                r.get("push_error", ""),
            ]
        )

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["G"].width = 42
    ws.column_dimensions["J"].width = 28
    ws.column_dimensions["K"].width = 40
    ws.column_dimensions["L"].width = 28

    wb.save(path)


def run_btvn_job(params: BtvnJobParams) -> tuple[bool, str, Path | None]:
    token = params.rikkei_token
    if not token.strip():
        return False, "Thiếu token Rikkei.", None

    try:
        from cham_bai import settings

        settings.api_key()
    except RuntimeError as e:
        return False, str(e), None

    try:
        students = fetch_students(token, params.class_id, params.session_id)
    except httpx.HTTPStatusError as e:
        return False, f"Lỗi API danh sách học sinh: {e.response.status_code} {e.response.text[:500]}", None
    except Exception as e:
        return False, f"Lỗi khi lấy danh sách học sinh: {e}", None

    if params.max_students is not None and params.max_students > 0:
        students = students[: params.max_students]

    all_rows: list[dict[str, Any]] = []

    for st in students:
        sid = _as_int(st.get("id"))
        if sid is None:
            continue
        code = str(st.get("studentCode", "") or "")
        name = str(st.get("fullName", "") or "")

        try:
            exercises = fetch_all_exercises_for_student(
                token,
                class_id=params.class_id.strip(),
                course_id=(params.course_id or "").strip(),
                session_id=params.session_id.strip(),
                student_id=sid,
            )
        except httpx.HTTPStatusError as e:
            return False, f"Lỗi API exercise (student {sid}): {e.response.status_code}", None
        except Exception as e:
            return False, f"Lỗi khi lấy bài tập học sinh {sid}: {e}", None

        for ex in exercises:
            eid = _as_int(ex.get("id"))
            if eid is None:
                continue
            link = ex.get("link_git") or ex.get("linkGit") or ""
            if isinstance(link, str):
                link = link.strip()
            else:
                link = ""

            hid = _homework_id(ex)
            title = _homework_title(ex)
            repo_key = _repo_group_key(link)
            mindmap = is_mindmap_homework_exercise(ex)

            link_valid = True
            link_reason = ""
            if not link:
                link_valid = False
                link_reason = "Chưa có link Git."
            elif repo_key is None:
                link_valid = False
                link_reason = "Link không phải GitHub hợp lệ (chỉ chấp nhận github.com / git@github.com)."

            all_rows.append(
                {
                    "student_id": sid,
                    "student_code": code,
                    "full_name": name,
                    "exercise_id": eid,
                    "homework_id": hid,
                    "homework_title": title,
                    "link_git": link,
                    "link_valid": link_valid,
                    "link_reason": link_reason,
                    "_repo_key": repo_key,
                    "mindmap_skip": mindmap,
                    "ai_comment": MINDMAP_PLACEHOLDER_COMMENT if mindmap else "",
                    "pushed": False,
                    "push_error": "",
                    "ai_error": "",
                    "repo_error": "",
                    "_raw_exercise": ex,
                }
            )

    mark_shared_repo_violations(all_rows)

    model = (params.text_model or "").strip() or os.environ.get(
        "OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6"
    )

    for r in all_rows:
        if not r.get("link_valid"):
            r["ai_comment"] = ""
            continue
        if r.get("mindmap_skip"):
            continue
        bundle = None
        try:
            bundle, err = fetch_repo_sources_bundle(str(r.get("link_git", "")))
            if err:
                r["repo_error"] = err
                bundle = None
        except Exception as e:
            r["repo_error"] = str(e)[:500]
            bundle = None
        try:
            r["ai_comment"] = generate_btvn_comment(
                homework_title=str(r.get("homework_title", "")),
                link_git=str(r.get("link_git", "")),
                model=model,
                submission_bundle=bundle,
            )
        except Exception as e:
            r["ai_comment"] = ""
            r["ai_error"] = str(e)[:800]
        time.sleep(max(0.0, params.openrouter_delay_s))

    if params.push_to_portal:
        for r in all_rows:
            if not r.get("link_valid"):
                continue
            if r.get("mindmap_skip"):
                comment = MINDMAP_PLACEHOLDER_COMMENT
            else:
                comment = (r.get("ai_comment") or "").strip()
            if not comment:
                r["push_error"] = "Không có nhận xét để ghi."
                continue
            eid = r["exercise_id"]
            detail = get_exercise_detail(token, eid)
            ok, err = put_exercise_comment(
                token,
                eid,
                comment=comment,
                link_git=str(r.get("link_git", "")),
                homework_id=r.get("homework_id") if isinstance(r.get("homework_id"), int) else None,
                full_body=detail,
            )
            r["pushed"] = ok
            r["push_error"] = err if not ok else ""

    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    out = Path(tmp_path)
    for r in all_rows:
        r.pop("_repo_key", None)
        r.pop("_raw_exercise", None)
    try:
        write_btvn_excel(all_rows, out)
    except Exception as e:
        return False, f"Lỗi ghi Excel: {e}", None

    return True, "", out
