"""
Giao diện web (FastAPI) cho Agent Edu — chạy local, tái dùng logic GUI.
"""

from __future__ import annotations

import asyncio
import json
import os
import re
import html as _html
import hashlib
import shutil
import tempfile
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import httpx
from fastapi import BackgroundTasks, FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from cham_bai import __version__
from cham_bai.gdocs_reader import fetch_google_doc_plain_text, is_google_docs_url
from cham_bai.git_remote import normalize_github_repo_url
from cham_bai.model_options import (
    DEFAULT_BTVN_MODEL,
    IMAGE_MODEL_OPTIONS,
    MODEL_OPTIONS,
    QUIZ_KIND_OPTIONS,
)
from cham_bai.quiz_excel import ensure_default_quiz_template, ensure_lesson_quiz_example_template
from cham_bai.quiz_excel import ensure_session_warmup_quiz_example_template
from cham_bai.quiz_gen import (
    QUIZ_KIND_LESSON,
    QUIZ_KIND_SESSION,
    QUIZ_KIND_SESSION_END,
    QUIZ_KIND_SESSION_WARMUP,
    QuizGenParams,
    default_quiz_output_path,
    default_session_end_quiz_output_path,
    default_session_warmup_quiz_output_path,
    normalize_quiz_kind,
    run_quiz_generation,
)
from cham_bai.reading_gen import (
    DEFAULT_LEARNING_GOALS,
    ReadingDocParams,
    reading_output_stem,
    run_reading_generation,
    sanitize_reading_filename_part,
)
from cham_bai.btvn_comment import BtvnCommentParams, run_btvn_comments_json
from cham_bai.rikkei_homework import (
    fetch_students as _btvn_fetch_students,
    fetch_all_exercises_for_student as _btvn_fetch_all_exercises_for_student,
    _homework_id as _btvn_homework_id,
)
from cham_bai.group_activity import GroupGradeParams, grade_group_activity
from cham_bai.hackathon_exam import (
    HackathonExamParams,
    build_hackathon_exam_docx_bytes,
    build_hackathon_exam_docx_from_spec,
)
from cham_bai.hackathon_ai import (
    build_required_bullets,
    ensure_practice_section,
    ensure_required_section,
    generate_hackathon_exam_spec,
    normalize_exam_template,
)
from cham_bai.workflow import (
    GradeJobParams,
    GradeJobResult,
    has_grade_slots,
    is_valid_report_source_url,
    run_grade_batch,
    run_grade_job,
)

_executor = ThreadPoolExecutor(max_workers=2, thread_name_prefix="web")


def _parse_bool_form(s: str | None) -> bool:
    return (s or "").strip().lower() in ("true", "1", "on", "yes")


def _static_dir() -> Path:
    return Path(__file__).resolve().parent / "web" / "static"


app = FastAPI(title="Agent Edu", version=__version__, docs_url=None, redoc_url=None)

if _static_dir().is_dir():
    app.mount("/static", StaticFiles(directory=str(_static_dir())), name="static")


@app.get("/", response_class=HTMLResponse)
async def index() -> HTMLResponse:
    index_path = _static_dir() / "index.html"
    if not index_path.is_file():
        return HTMLResponse(
            "<h1>Agent Edu web</h1><p>Thiếu file static/index.html</p>",
            status_code=500,
        )
    return HTMLResponse(index_path.read_text(encoding="utf-8"))


@app.get("/api/meta")
async def api_meta() -> JSONResponse:
    return JSONResponse(
        {
            "version": __version__,
            "models": list(MODEL_OPTIONS),
            "image_models": list(IMAGE_MODEL_OPTIONS),
            "quiz_kinds": [{"label": a, "value": b} for a, b in QUIZ_KIND_OPTIONS],
            "default_model": os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6"),
            "default_btvn_model": os.getenv("OPENROUTER_BTVN_MODEL", DEFAULT_BTVN_MODEL),
            "default_image_model": IMAGE_MODEL_OPTIONS[0],
            "default_learning_goals": DEFAULT_LEARNING_GOALS,
        }
    )


def _format_grade_block(label: str, res: GradeJobResult) -> str:
    lines = [f"——— {label} ———"]
    for w in res.warnings:
        lines.append(f"[Cảnh báo] {w}")
    if not res.ok:
        lines.append(f"[Lỗi] {res.error_message or '?'}")
        return "\n".join(lines) + "\n"
    lines.append("[Xong]")
    if res.json_text:
        try:
            blob = json.loads(res.json_text)
            lines.append(f"Điểm: {blob.get('final_score')}")
            lines.append(f"Nhận xét: {blob.get('final_comment')}")
            mp = blob.get("mini_project_present")
            if mp is not None:
                lines.append(f"Mini project (chỉ có/không): {mp}")
        except json.JSONDecodeError:
            lines.append(res.json_text[:2000])
    return "\n".join(lines) + "\n"


async def _save_upload_optional(upload: UploadFile | None, suffix: str) -> str | None:
    if upload is None or not upload.filename:
        return None
    data = await upload.read()
    if not data:
        return None
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    Path(path).write_bytes(data)
    return path


def _run_grade_sync(
    assignment_ref: str,
    subs_lines: list[str],
    model: str,
    no_template: bool,
    strict_ai: bool,
    ai_confidence: int,
    report_repos_text: str = "",
) -> tuple[bool, str, list[dict]]:
    try:
        from cham_bai.settings import api_key as _need_key

        _need_key()
    except RuntimeError as e:
        return False, str(e), []

    params = GradeJobParams(
        assignment_ref=assignment_ref,
        submission_ref=subs_lines[0],
        out_path=None,
        model=model.strip(),
        no_template=no_template,
        strict_ai=strict_ai,
        ai_confidence=int(ai_confidence),
        debug=False,
        report_repos_text=report_repos_text or "",
    )

    log_parts: list[str] = []
    payload: list[dict] = []

    if len(subs_lines) == 1:
        res = run_grade_job(params)
        log_parts.append(_format_grade_block("Bài nộp", res))
        row: dict = {"submission": subs_lines[0], "ok": res.ok}
        if res.error_message:
            row["error"] = res.error_message
        if res.warnings:
            row["warnings"] = res.warnings
        if res.json_text:
            try:
                row["result"] = json.loads(res.json_text)
            except json.JSONDecodeError:
                row["result_raw"] = res.json_text
        payload.append(row)
        return res.ok, "\n".join(log_parts), payload

    batch = run_grade_batch(assignment_ref, subs_lines, params)
    ok_any = False
    for sub, res in batch:
        log_parts.append(_format_grade_block(sub, res))
        row = {"submission": sub, "ok": res.ok}
        if res.error_message:
            row["error"] = res.error_message
        if res.warnings:
            row["warnings"] = res.warnings
        if res.json_text:
            try:
                row["result"] = json.loads(res.json_text)
            except json.JSONDecodeError:
                row["result_raw"] = res.json_text
        payload.append(row)
        if res.ok:
            ok_any = True
    return ok_any, "\n".join(log_parts), payload


@app.post("/api/grade")
async def api_grade(
    assignment_text: str = Form(...),
    submissions_text: str = Form(""),
    report_repos_text: str = Form(""),
    model: str = Form(""),
    use_template: str = Form("true"),
    strict_ai: str = Form("true"),
    ai_confidence: int = Form(75),
) -> JSONResponse:
    subs_lines = [(ln.rstrip("\r") or "").strip() for ln in (submissions_text or "").splitlines()]

    assignment_ref = (assignment_text or "").strip()
    if not assignment_ref:
        raise HTTPException(
            status_code=400,
            detail="Điền đề bài: đường dẫn file .docx trên máy chủ hoặc link Google Docs.",
        )

    p = Path(assignment_ref)
    ok_docx = p.is_file() and p.suffix.lower() == ".docx"
    ok_gdoc = is_google_docs_url(assignment_ref)
    if not ok_docx and not ok_gdoc:
        raise HTTPException(
            status_code=400,
            detail="Đề cần là đường dẫn file .docx tồn tại trên máy chủ hoặc URL Google Docs.",
        )

    if not has_grade_slots(subs_lines, report_repos_text or ""):
        raise HTTPException(
            status_code=400,
            detail="Cần ít nhất một dòng có bài nộp (thư mục/GitHub) hoặc link repo báo cáo.",
        )

    for s in subs_lines:
        if not s.strip():
            continue
        if not normalize_github_repo_url(s) and not Path(s).is_dir():
            raise HTTPException(
                status_code=400,
                detail=f"Bài nộp không hợp lệ (thư mục trên máy chủ hoặc link GitHub): {s[:160]}",
            )

    for ln in (report_repos_text or "").splitlines():
        u = ln.strip()
        if u and not is_valid_report_source_url(u):
            raise HTTPException(
                status_code=400,
                detail=f"Dòng báo cáo không hợp lệ (GitHub, Google Docs hoặc OneDrive/SharePoint): {u[:160]}",
            )

    loop = asyncio.get_event_loop()

    def work():
        return _run_grade_sync(
            assignment_ref,
            subs_lines,
            model or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6"),
            not _parse_bool_form(use_template),
            _parse_bool_form(strict_ai),
            ai_confidence,
            report_repos_text=report_repos_text or "",
        )

    ok, log, results = await loop.run_in_executor(_executor, work)

    return JSONResponse({"ok": ok, "log": log, "results": results})


@app.post("/api/quiz")
async def api_quiz(
    background_tasks: BackgroundTasks,
    quiz_kind_label: str = Form(...),
    subject: str = Form(""),
    lesson: str = Form(""),
    session: str = Form(""),
    session_prev: str = Form(""),
    session_current: str = Form(""),
    lecture_gdocs_urls_text: str = Form(""),
    lecture_prev_gdocs_urls_text: str = Form(""),
    lecture_current_gdocs_urls_text: str = Form(""),
    num_questions: int = Form(5),
    model: str = Form(""),
    template_file: UploadFile | None = File(None),
    docx_file: UploadFile | None = File(None),
) -> FileResponse:
    kind_map = dict(QUIZ_KIND_OPTIONS)
    qkind = normalize_quiz_kind(kind_map.get(quiz_kind_label.strip(), QUIZ_KIND_SESSION))

    tmp_tpl = await _save_upload_optional(template_file, ".xlsx")
    if tmp_tpl:
        template_path = Path(tmp_tpl)
    elif qkind == QUIZ_KIND_LESSON:
        template_path = ensure_lesson_quiz_example_template()
    elif qkind in (QUIZ_KIND_SESSION_WARMUP, QUIZ_KIND_SESSION_END):
        template_path = ensure_session_warmup_quiz_example_template()
    else:
        template_path = ensure_default_quiz_template()

    # Lecture source (ưu tiên Google Docs nhiều link; fallback DOCX upload).
    # Allow long URLs to wrap/break into multiple lines: join continuation lines.
    def _parse_docs_urls(text: str) -> list[str]:
        out: list[str] = []
        buf2 = ""
        for ln in (text or "").splitlines():
            s = (ln or "").strip()
            if not s or s.startswith("#"):
                continue
            if s.lower().startswith(("http://", "https://")):
                if buf2:
                    out.append(buf2)
                buf2 = s
            else:
                buf2 = (buf2 + s) if buf2 else s
        if buf2:
            out.append(buf2)
        return out

    def _fetch_docs_text(urls: list[str], *, label: str) -> str:
        if not urls:
            return ""
        parts: list[str] = []
        for idx, u in enumerate(urls[:12], start=1):
            if not is_google_docs_url(u):
                raise HTTPException(status_code=400, detail=f"Link Google Docs không hợp lệ ({label}, dòng {idx}): {u[:180]}")
            try:
                txt = fetch_google_doc_plain_text(u)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Lỗi đọc Google Docs ({label}, dòng {idx}): {e}")
            txt = (txt or "").strip()
            if txt:
                parts.append(f"=== {label} DOC {idx} ===\nNguồn: {u}\n\n{txt}\n")
        return ("\n\n".join(parts)).strip()

    lecture_text = ""
    lecture_prev_text = ""
    lecture_curr_text = ""
    try:
        # Warmup: 2 nhóm link (session cũ / session hiện tại)
        prev_urls = _parse_docs_urls(lecture_prev_gdocs_urls_text)
        curr_urls = _parse_docs_urls(lecture_current_gdocs_urls_text)
        lecture_prev_text = _fetch_docs_text(prev_urls, label="SESSION TRƯỚC") if prev_urls else ""
        lecture_curr_text = _fetch_docs_text(curr_urls, label="SESSION HIỆN TẠI") if curr_urls else ""

        # Fallback: ô 1 nhóm link (cũ)
        urls_single = _parse_docs_urls(lecture_gdocs_urls_text)
        lecture_text = _fetch_docs_text(urls_single, label="BÀI GIẢNG") if urls_single else ""
    except HTTPException:
        _cleanup_quiz_temp(tmp_tpl, None, None)
        raise

    tmp_docx = await _save_upload_optional(docx_file, ".docx")
    docx_path = Path(tmp_docx) if tmp_docx else None
    if docx_path and not docx_path.is_file():
        if tmp_tpl:
            try:
                os.unlink(tmp_tpl)
            except OSError:
                pass
        if tmp_docx:
            try:
                os.unlink(tmp_docx)
            except OSError:
                pass
        raise HTTPException(status_code=400, detail="File DOCX không hợp lệ.")

    out_dir = tempfile.mkdtemp(prefix="quiz_out_")
    lesson_s = (lesson or "").strip()
    session_s = (session or "").strip()
    subj_s = (subject or "").strip()
    prev_s = (session_prev or "").strip()
    curr_s = (session_current or "").strip()
    if qkind == QUIZ_KIND_SESSION_WARMUP:
        if not subj_s or not curr_s:
            _cleanup_quiz_temp(tmp_tpl, tmp_docx, out_dir)
            raise HTTPException(status_code=400, detail="Thiếu môn học hoặc session hiện tại.")
        lesson_s = subj_s
        session_s = curr_s
        out_path = Path(out_dir) / default_session_warmup_quiz_output_path(template_path, curr_s).name
    elif qkind == QUIZ_KIND_SESSION_END:
        if not subj_s or not curr_s:
            _cleanup_quiz_temp(tmp_tpl, tmp_docx, out_dir)
            raise HTTPException(status_code=400, detail="Thiếu môn học hoặc session hiện tại.")
        lesson_s = subj_s
        session_s = curr_s
        out_path = Path(out_dir) / default_session_end_quiz_output_path(template_path, curr_s).name
    else:
        out_path = Path(out_dir) / default_quiz_output_path(template_path, lesson_s, session_s).name

    try:
        from cham_bai.settings import api_key as _need_key

        _need_key()
    except RuntimeError as e:
        _cleanup_quiz_temp(tmp_tpl, tmp_docx, out_dir)
        raise HTTPException(status_code=400, detail=str(e)) from e

    params = QuizGenParams(
        template_xlsx=template_path,
        docx_path=docx_path,
        lecture_text=lecture_text,
        lecture_text_prev=lecture_prev_text,
        lecture_text_current=lecture_curr_text,
        lesson=lesson_s,
        session=session_s,
        session_prev=prev_s,
        session_current=curr_s,
        num_questions=(45 if qkind in (QUIZ_KIND_SESSION_WARMUP, QUIZ_KIND_SESSION_END) else int(num_questions)),
        model=(model or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6")).strip(),
        output_xlsx=out_path,
        subject=subj_s,
        quiz_kind=qkind,
    )

    loop = asyncio.get_event_loop()

    def gen():
        return run_quiz_generation(params)

    ok, msg = await loop.run_in_executor(_executor, gen)
    if tmp_tpl:
        try:
            os.unlink(tmp_tpl)
        except OSError:
            pass
    if tmp_docx:
        try:
            os.unlink(tmp_docx)
        except OSError:
            pass

    if not ok:
        shutil.rmtree(out_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=msg)

    safe_name = out_path.name
    background_tasks.add_task(shutil.rmtree, out_dir, True)

    return FileResponse(
        path=str(out_path),
        filename=safe_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _cleanup_quiz_temp(tmp_tpl: str | None, tmp_docx: str | None, out_dir: str | None) -> None:
    if tmp_tpl:
        try:
            os.unlink(tmp_tpl)
        except OSError:
            pass
    if tmp_docx:
        try:
            os.unlink(tmp_docx)
        except OSError:
            pass
    if out_dir:
        shutil.rmtree(out_dir, ignore_errors=True)


def _rk_bearer(token: str) -> str:
    t = (token or "").strip()
    if not t:
        return ""
    if t.lower().startswith("bearer "):
        return t
    return "Bearer " + t


def _rk_sanitize_html(s: str) -> str:
    raw = (s or "").strip()
    # drop script/style and basic inline handlers
    raw = re.sub(r"<script[\s\S]*?</script>", "", raw, flags=re.I)
    raw = re.sub(r"<style[\s\S]*?</style>", "", raw, flags=re.I)
    raw = re.sub(r"\son\w+=(\"[^\"]*\"|'[^']*')", "", raw, flags=re.I)
    return raw


_IMG_SRC_RE = re.compile(r"<img[^>]+src=[\"'](?P<src>[^\"']+)[\"'][^>]*>", re.I)


def _rk_extract_image_urls(html: str) -> list[str]:
    seen: set[str] = set()
    urls: list[str] = []
    for m in _IMG_SRC_RE.finditer(html or ""):
        u = (m.group("src") or "").strip()
        if not u:
            continue
        if u not in seen:
            seen.add(u)
            urls.append(u)
    return urls


def _rk_html_to_plain_text(html: str) -> str:
    s = (html or "").strip()
    if not s:
        return ""
    # line breaks for common tags
    s = re.sub(r"(?i)<br\s*/?>", "\n", s)
    s = re.sub(r"(?i)</p\s*>", "\n\n", s)
    s = re.sub(r"(?i)</li\s*>", "\n", s)
    # remove tags
    s = re.sub(r"<[^>]+>", " ", s)
    s = _html.unescape(s)
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    s = re.sub(r"\n\s+\n", "\n\n", s)
    return s.strip()


async def _rk_fetch_image_bytes(url: str) -> tuple[str, bytes] | None:
    u = (url or "").strip()
    if not u.startswith(("http://", "https://")):
        return None
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(u, headers={"User-Agent": "AgentEdu/1.0"})
        r.raise_for_status()
        raw = r.content or b""
        if not raw or len(raw) < 64:
            return None
        if len(raw) > 5_000_000:
            return None
        ct = (r.headers.get("content-type") or "").split(";")[0].strip().lower()
        if ct and ct.startswith("image/"):
            return ct, raw
        # best-effort sniff
        if raw[:8] == b"\x89PNG\r\n\x1a\n":
            return "image/png", raw
        if raw[:2] == b"\xff\xd8":
            return "image/jpeg", raw
        if raw[:6] in (b"GIF87a", b"GIF89a"):
            return "image/gif", raw
        if raw[:4] == b"RIFF" and raw[8:12] == b"WEBP":
            return "image/webp", raw
        return None
    except Exception:
        return None


@app.post("/api/rikkei/session")
async def api_rikkei_session(
    session_id: int = Form(...),
    rikkei_token: str = Form(...),
) -> JSONResponse:
    sid = int(session_id)
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    url = f"https://apiportal.rikkei.edu.vn/sessions/{sid}"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền truy cập session này.")
        r.raise_for_status()
        data = r.json()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei: {e}")

    hw = data.get("homework") if isinstance(data, dict) else None
    if not isinstance(hw, list):
        hw = []
    out_hw: list[dict] = []
    for item in hw:
        if not isinstance(item, dict):
            continue
        hid = item.get("id")
        title = item.get("title") or ""
        desc = item.get("description") or ""
        desc_s = _rk_sanitize_html(str(desc))
        out_hw.append(
            {
                "id": hid,
                "title": str(title),
                "description_html": desc_s,
                "plain_text": _rk_html_to_plain_text(desc_s),
                "image_urls": _rk_extract_image_urls(desc_s),
            }
        )

    return JSONResponse(
        {
            "id": data.get("id") if isinstance(data, dict) else sid,
            "name": data.get("name") if isinstance(data, dict) else f"Session {sid}",
            "homework": out_hw,
        }
    )


@app.post("/api/rikkei/btvn/students")
async def api_rikkei_btvn_students(
    rikkei_token: str = Form(...),
    class_id: str = Form(...),
    session_id: str = Form(...),
) -> JSONResponse:
    try:
        items = _btvn_fetch_students(rikkei_token, class_id, session_id)
    except httpx.HTTPStatusError as e:
        raise HTTPException(
            status_code=502,
            detail=f"Lỗi API danh sách học sinh: {e.response.status_code} {e.response.text[:500]}",
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi khi lấy danh sách học sinh: {e}")
    return JSONResponse({"ok": True, "items": items})


def _extract_rikkei_token(payload: object) -> str:
    """
    Best-effort: Rikkei portal implementations vary.
    Try common shapes: {"token":..}, {"accessToken":..}, {"data":{"token":..}}, etc.
    """
    if isinstance(payload, dict):
        for k in ("token", "accessToken", "access_token", "secretToken", "secret_token", "jwt"):
            v = payload.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip()
        data = payload.get("data")
        if isinstance(data, dict):
            for k in ("token", "accessToken", "access_token", "secretToken", "secret_token", "jwt"):
                v = data.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip()
    return ""


def _unwrap_list_payload(payload: object) -> list[dict]:
    if isinstance(payload, list):
        return [x for x in payload if isinstance(x, dict)]
    if isinstance(payload, dict):
        for k in ("data", "items", "content", "records", "result", "rows", "resultTest", "result_test"):
            v = payload.get(k)
            if isinstance(v, list):
                return [x for x in v if isinstance(x, dict)]
            # Common pattern: {"data": {"data": [...]}} or {"data": {"items": [...]}}
            if isinstance(v, dict):
                for k2 in ("data", "items", "content", "records", "result", "rows"):
                    v2 = v.get(k2)
                    if isinstance(v2, list):
                        return [x for x in v2 if isinstance(x, dict)]
    return []


def _pick_first_str(d: dict, keys: list[str]) -> str:
    for k in keys:
        v = d.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    # case-insensitive fallback
    lower = {str(k).lower(): k for k in d.keys()}
    for k in keys:
        k2 = lower.get(k.lower())
        if k2 is None:
            continue
        v = d.get(k2)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _pick_first_int(d: dict, keys: list[str]) -> int | None:
    for k in keys:
        v = d.get(k)
        if isinstance(v, int):
            return v
        if isinstance(v, str) and v.strip().isdigit():
            return int(v.strip())
    lower = {str(k).lower(): k for k in d.keys()}
    for k in keys:
        k2 = lower.get(k.lower())
        if k2 is None:
            continue
        v = d.get(k2)
        if isinstance(v, int):
            return v
        if isinstance(v, str) and v.strip().isdigit():
            return int(v.strip())
    return None


def _norm_system_item(x: dict) -> dict:
    return {
        "id": _pick_first_int(x, ["id", "systemId", "system_id"]) or x.get("id"),
        "systemCode": _pick_first_str(x, ["systemCode", "system_code", "code", "system_code_name"]),
        "name": _pick_first_str(x, ["name", "systemName", "system_name", "title"]),
        "_raw": x,
    }


def _norm_course_item(x: dict) -> dict:
    return {
        "id": _pick_first_int(x, ["id", "courseId", "course_id"]) or x.get("id"),
        "courseCode": _pick_first_str(x, ["courseCode", "course_code", "code"]),
        "name": _pick_first_str(x, ["name", "courseName", "course_name", "title"]),
        "_raw": x,
    }


def _norm_class_item(x: dict) -> dict:
    return {
        "id": _pick_first_int(x, ["id", "classId", "class_id"]) or x.get("id"),
        "classCode": _pick_first_str(x, ["classCode", "class_code", "code"]),
        "name": _pick_first_str(x, ["name", "className", "class_name", "title"]),
        "_raw": x,
    }


def _norm_session_item(x: dict) -> dict:
    return {
        "id": _pick_first_int(x, ["id", "sessionId", "session_id"]) or x.get("id"),
        "name": _pick_first_str(x, ["name", "title"]),
        "position": _pick_first_int(x, ["position", "stt", "order"]),
        "type": _pick_first_str(x, ["type", "sessionType", "session_type"]),
        "miniProject": _pick_first_str(x, ["miniProject", "mini_project", "miniproject"]),
        "_raw": x,
    }


def _norm_practice_resource_item(x: dict) -> dict:
    st = x.get("student") if isinstance(x, dict) else None
    if not isinstance(st, dict):
        st = {}
    return {
        "id": _pick_first_int(x, ["id"]) or x.get("id"),
        "link": _pick_first_str(x, ["link"]),
        "reportLink": _pick_first_str(x, ["reportLink", "report_link"]),
        "score": _pick_first_int(x, ["score"]),
        "comment": _pick_first_str(x, ["comment"]),
        "studentId": _pick_first_int(st, ["id", "studentId", "student_id"]),
        "studentCode": _pick_first_str(st, ["studentCode", "student_code", "code"]),
        "fullName": _pick_first_str(st, ["fullName", "full_name", "name"]),
        "_raw": x,
    }


@app.post("/api/rikkei/login")
async def api_rikkei_login(
    email: str = Form(...),
    password: str = Form(...),
) -> JSONResponse:
    em = (email or "").strip()
    pw = password or ""
    if not em or not pw:
        raise HTTPException(status_code=400, detail="Thiếu email hoặc mật khẩu.")

    url = "https://apiportal.rikkei.edu.vn/auth/secret-token"
    # Do NOT log credentials. Only exchange to token.
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            # Attempt 1: BasicAuth, empty JSON (some setups use only auth header)
            r = await client.post(url, auth=(em, pw), headers={"User-Agent": "AgentEdu/1.0"})
            if r.status_code >= 400:
                # Attempt 2: JSON body with email/password
                r = await client.post(
                    url,
                    json={"email": em, "password": pw},
                    headers={"User-Agent": "AgentEdu/1.0", "Accept": "application/json"},
                )
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Sai tài khoản hoặc không có quyền.")
        r.raise_for_status()
        data = r.json()
        tok = _extract_rikkei_token(data)
        if not tok:
            # fallback: if API returns raw string
            if isinstance(data, str) and data.strip():
                tok = data.strip()
        if not tok:
            raise HTTPException(status_code=502, detail="Đăng nhập OK nhưng không tìm thấy token trong response.")
        return JSONResponse({"ok": True, "token": tok})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi đăng nhập Rikkei: {e}")


@app.post("/api/rikkei/systems")
async def api_rikkei_systems(
    rikkei_token: str = Form(...),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    url = "https://apiportal.rikkei.edu.vn/automation/systems"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_system_item(x) for x in raw_items if isinstance(x, dict)]
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (systems): {e}")


@app.post("/api/rikkei/classes")
async def api_rikkei_classes(
    rikkei_token: str = Form(...),
    system_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    sid = (system_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not sid:
        raise HTTPException(status_code=400, detail="Thiếu system_id.")
    url = f"https://apiportal.rikkei.edu.vn/automation/systems/{sid}/classes"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_class_item(x) for x in raw_items if isinstance(x, dict)]
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (classes): {e}")


@app.post("/api/rikkei/class-courses")
async def api_rikkei_class_courses(
    rikkei_token: str = Form(...),
    class_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    cid = (class_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not cid:
        raise HTTPException(status_code=400, detail="Thiếu class_id.")
    url = f"https://apiportal.rikkei.edu.vn/automation/classes/{cid}/courses"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_course_item(x) for x in raw_items if isinstance(x, dict)]
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (class-courses): {e}")


@app.post("/api/rikkei/course-sessions")
async def api_rikkei_course_sessions(
    rikkei_token: str = Form(...),
    course_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    cid = (course_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not cid:
        raise HTTPException(status_code=400, detail="Thiếu course_id.")
    url = f"https://apiportal.rikkei.edu.vn/sessions/course/{cid}"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items_all = [_norm_session_item(x) for x in raw_items if isinstance(x, dict)]
        # Return all sessions (not only THỰC HÀNH)
        items = items_all
        items.sort(key=lambda z: int(z.get("position") or 10_000))
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (course-sessions): {e}")


@app.post("/api/rikkei/practice-resource")
async def api_rikkei_practice_resource(
    rikkei_token: str = Form(...),
    class_id: str = Form(""),
    session_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    cid = (class_id or "").strip()
    sid = (session_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not cid or not sid:
        raise HTTPException(status_code=400, detail="Thiếu class_id hoặc session_id.")
    url = "https://apiportal.rikkei.edu.vn/practice-resource"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(
                url,
                params={"classId": cid, "sessionId": sid},
                headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"},
            )
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_practice_resource_item(x) for x in raw_items if isinstance(x, dict)]
        # sort by studentCode then name for stable display
        items.sort(
            key=lambda z: (
                str(z.get("studentCode") or ""),
                str(z.get("fullName") or ""),
                int(z.get("id") or 10_000_000),
            )
        )
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (practice-resource): {e}")


def _looks_html(s: str) -> bool:
    t = (s or "").strip()
    return bool(re.search(r"<\s*(p|br|div|span|ul|li)\b", t, re.I))


@app.post("/api/rikkei/practice-resource/patch-batch")
async def api_rikkei_practice_resource_patch_batch(
    rikkei_token: str = Form(...),
    patches_json: str = Form(...),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    try:
        patches = json.loads(patches_json or "[]")
    except Exception:
        raise HTTPException(status_code=400, detail="patches_json không phải JSON hợp lệ.")
    if not isinstance(patches, list) or not patches:
        raise HTTPException(status_code=400, detail="Danh sách patch rỗng.")

    ok_count = 0
    fails: list[dict] = []
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        for p in patches[:500]:
            if not isinstance(p, dict):
                continue
            pid = p.get("id")
            score = p.get("score")
            comment_html = str(p.get("comment_html") or "").strip()
            try:
                pid_int = int(pid)
            except Exception:
                fails.append({"id": pid, "error": "id không hợp lệ"})
                continue
            try:
                score_int = int(score)
            except Exception:
                fails.append({"id": pid_int, "error": "score không hợp lệ"})
                continue
            body = {"score": score_int}
            if comment_html:
                body["comment"] = comment_html if _looks_html(comment_html) else f"<p>{_html.escape(comment_html)}</p>"
            url = f"https://apiportal.rikkei.edu.vn/practice-resource/{pid_int}"
            try:
                r = await client.patch(
                    url,
                    json=body,
                    headers={
                        "Authorization": _rk_bearer(tok),
                        "User-Agent": "AgentEdu/1.0",
                        "Accept": "application/json",
                    },
                )
                if r.status_code in (401, 403):
                    fails.append({"id": pid_int, "error": "Token không hợp lệ hoặc không có quyền"})
                    continue
                r.raise_for_status()
                ok_count += 1
            except Exception as e:
                fails.append({"id": pid_int, "error": str(e)[:240]})

    return JSONResponse(
        {
            "ok": True,
            "ok_count": ok_count,
            "fail_count": len(fails),
            "fails": fails[:50],
        }
    )


def _norm_test_schedule_item(x: dict) -> dict:
    test = x.get("test") if isinstance(x.get("test"), dict) else {}
    cls = x.get("class") if isinstance(x.get("class"), dict) else {}
    return {
        "id": _pick_first_int(x, ["id"]) or x.get("id"),
        "type": _pick_first_str(x, ["type"]),
        "testId": _pick_first_int(test, ["id"]) or test.get("id"),
        "testName": _pick_first_str(test, ["testName", "name", "title"]),
        "classCode": _pick_first_str(cls, ["classCode", "class_code", "code"]),
        "className": _pick_first_str(cls, ["name", "className"]),
        "testStart": _pick_first_str(x, ["testStart", "start"]),
        "testEnd": _pick_first_str(x, ["testEnd", "end"]),
        "_raw": x,
    }


def _extract_exam_docs_from_html(html: str) -> dict[str, str]:
    """
    Return mapping {"01": url, "02": url, ...} from test question HTML.
    Looks for "Đề 01" around google docs links.
    """
    s = html or ""
    out: dict[str, str] = {}
    # find all anchor hrefs to google docs
    for m in re.finditer(r'href="(?P<u>https?://docs\.google\.com/document/d/[^"]+)"', s, re.I):
        u = (m.group("u") or "").strip()
        # look back window for "Đề xx"
        start = max(0, m.start() - 120)
        chunk = s[start : m.start()]
        m2 = re.search(r"Đề\s*0?([0-9]{1,3})", chunk, re.I)
        if m2:
            k = int(m2.group(1))
            if 1 <= k <= 99:
                out[f"{k:02d}"] = u
    return out


def _norm_test_detail(data: dict) -> dict:
    qts = data.get("questionTests") if isinstance(data.get("questionTests"), list) else []
    docs: dict[str, str] = {}
    for q in qts:
        if not isinstance(q, dict):
            continue
        content = str(q.get("content") or "")
        docs.update(_extract_exam_docs_from_html(content))
    return {"id": data.get("id"), "testName": data.get("testName"), "docs": docs}


def _norm_test_schedule_detail_item(x: dict) -> dict:
    st = x.get("student") if isinstance(x.get("student"), dict) else {}
    return {
        "id": _pick_first_int(x, ["id"]) or x.get("id"),
        "point": x.get("point"),
        "link": _pick_first_str(x, ["link"]),
        "submittedAt": _pick_first_str(x, ["submittedAt", "submitted_at"]),
        "studentCode": _pick_first_str(st, ["studentCode", "student_code", "code"]),
        "fullName": _pick_first_str(st, ["fullName", "full_name", "name"]),
        "_raw": x,
    }


def _extract_exam_code_from_text(s: str) -> int | None:
    t = str(s or "").strip()
    if not t:
        return None
    m = re.search(r"(?:[_-])0*([0-9]{1,3})$", t, re.I)
    if not m:
        # Hỗ trợ dạng: CSDL_..._004_hackthon (mã nằm giữa chuỗi).
        m = re.search(r"(?:[_-])0*([0-9]{1,3})(?:[_-][a-z0-9].*)?$", t, re.I)
    if not m:
        # Hỗ trợ dạng: ...-004- (ký tự phân tách treo cuối chuỗi).
        m = re.search(r"(?:[_-])0*([0-9]{1,3})[_-]+$", t, re.I)
    if not m:
        m = re.search(r"([0-9]{1,3})$", t, re.I)
    if not m:
        return None
    try:
        n = int(m.group(1))
    except Exception:
        return None
    return n if 1 <= n <= 999 else None


@app.post("/api/github/exam-code")
async def api_github_exam_code(repo_url: str = Form(...)) -> JSONResponse:
    u = (repo_url or "").strip()
    norm = normalize_github_repo_url(u) or ""
    if not norm:
        raise HTTPException(status_code=400, detail="repo_url không hợp lệ.")
    m = re.match(r"^https?://github\.com/([^/]+)/([^/?#]+)", norm, re.I)
    if not m:
        raise HTTPException(status_code=400, detail="repo_url không phải link GitHub owner/repo hợp lệ.")
    owner = m.group(1).strip()
    repo = m.group(2).strip().replace(".git", "")

    # 1) lưu mã suy đoán từ tên repo, nhưng KHÔNG trả ngay:
    # ưu tiên mã thật lấy từ tên file/folder trong repo (ổn định hơn với hậu tố kiểu 15h28).
    repo_guess = _extract_exam_code_from_text(repo)

    gh_url = f"https://api.github.com/repos/{owner}/{repo}/contents"
    headers = {"User-Agent": "AgentEdu/1.0", "Accept": "application/vnd.github+json"}
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            r = await client.get(gh_url, headers=headers)
        if r.status_code == 404:
            return JSONResponse({"ok": True, "code": None, "source": "not_found"})
        r.raise_for_status()
        arr = r.json()
        if not isinstance(arr, list):
            return JSONResponse({"ok": True, "code": None, "source": "empty"})

        # 2) quét tên item ở root (file/folder)
        for it in arr:
            if not isinstance(it, dict):
                continue
            name = str(it.get("name") or "").strip()
            name = re.sub(r"\.(sql|py|js|ts|java|cpp|c|cs)$", "", name, flags=re.I)
            n2 = _extract_exam_code_from_text(name)
            if n2:
                return JSONResponse({"ok": True, "code": n2, "source": "root_item_name"})
        if repo_guess:
            return JSONResponse({"ok": True, "code": repo_guess, "source": "repo_name_guess"})
        return JSONResponse({"ok": True, "code": None, "source": "none"})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi đọc GitHub contents: {e}")


@app.post("/api/rikkei/test-schedules")
async def api_rikkei_test_schedules(
    rikkei_token: str = Form(...),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    url = "https://apiportal.rikkei.edu.vn/test-schedule/user/my-schedules/me"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_test_schedule_item(x) for x in raw_items if isinstance(x, dict)]
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (test-schedules): {e}")


@app.post("/api/rikkei/test")
async def api_rikkei_test(
    rikkei_token: str = Form(...),
    test_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    tid = (test_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not tid:
        raise HTTPException(status_code=400, detail="Thiếu test_id.")
    url = f"https://apiportal.rikkei.edu.vn/tests/{tid}"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        payload = r.json()
        d = payload.get("data") if isinstance(payload, dict) else None
        if not isinstance(d, dict):
            raise HTTPException(status_code=502, detail="Response tests/{id} không có field data.")
        out = _norm_test_detail(d)
        return JSONResponse({"ok": True, **out})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (test): {e}")


@app.post("/api/rikkei/test-schedule-detail")
async def api_rikkei_test_schedule_detail(
    rikkei_token: str = Form(...),
    schedule_id: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    sid = (schedule_id or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    if not sid:
        raise HTTPException(status_code=400, detail="Thiếu schedule_id.")
    url = f"https://apiportal.rikkei.edu.vn/test-schedule/detail/{sid}"
    try:
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
            r = await client.get(url, headers={"Authorization": _rk_bearer(tok), "User-Agent": "AgentEdu/1.0"})
        if r.status_code in (401, 403):
            raise HTTPException(status_code=401, detail="Token không hợp lệ hoặc không có quyền.")
        r.raise_for_status()
        raw_items = _unwrap_list_payload(r.json())
        items = [_norm_test_schedule_detail_item(x) for x in raw_items if isinstance(x, dict)]
        return JSONResponse({"ok": True, "items": items})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi gọi API Rikkei (test-schedule-detail): {e}")


def _result_test_point_str(point_f: float) -> str:
    """Rikkei portal nhận `point` dạng chuỗi (Hackathon thang 100)."""
    point_f = max(0.0, min(100.0, float(point_f)))
    if abs(point_f - int(point_f)) < 1e-9:
        return str(int(point_f))
    s = f"{point_f:.2f}".rstrip("0").rstrip(".")
    return s if s else "0"


RESULT_TEST_NOTE_MAX = 200


def _normalize_result_test_note(note: str) -> str:
    s = " ".join(str(note or "").split()).strip()
    if not s:
        return ""
    if len(s) <= RESULT_TEST_NOTE_MAX:
        out0 = s
        out0 = re.sub(r"[:;,]\s*$", ".", out0)
        out0 = re.sub(r"\b(tuy nhiên|nhưng|và|cụ thể|gồm)\s*[:;,]?\s*$", ".", out0, flags=re.I)
        out0 = re.sub(r"\s{2,}", " ", out0).strip()
        return out0

    # Ưu tiên cắt theo ranh giới câu để tránh bị cụt "... và", "... thay vì".
    clipped = s[:RESULT_TEST_NOTE_MAX]
    cuts = [clipped.rfind("."), clipped.rfind("!"), clipped.rfind("?"), clipped.rfind(";"), clipped.rfind(":")]
    cut_at = max(cuts)
    if cut_at >= 80:
        out = clipped[: cut_at + 1].strip()
    else:
        # fallback theo khoảng trắng gần cuối
        ws = clipped.rfind(" ")
        out = (clipped[:ws] if ws >= 60 else clipped).strip()
        if out and out[-1] not in ".!?":
            out += "."
    out = re.sub(r"[:;,]\s*$", ".", out)
    out = re.sub(r"\b(tuy nhiên|nhưng|và|cụ thể|gồm)\s*[:;,]?\s*$", ".", out, flags=re.I)
    out = re.sub(r"\s{2,}", " ", out).strip()
    return out


async def _rikkei_patch_result_test(
    client: httpx.AsyncClient,
    tok: str,
    pid_int: int,
    body: dict,
) -> httpx.Response:
    url = f"https://apiportal.rikkei.edu.vn/result-test/{pid_int}"
    return await client.patch(
        url,
        json=body,
        headers={
            "Authorization": _rk_bearer(tok),
            "User-Agent": "AgentEdu/1.0",
            "Accept": "application/json",
            "Content-Type": "application/json",
        },
    )


@app.post("/api/rikkei/result-test/patch-batch")
async def api_rikkei_result_test_patch_batch(
    rikkei_token: str = Form(...),
    patches_json: str = Form(...),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    try:
        patches = json.loads(patches_json or "[]")
    except Exception:
        raise HTTPException(status_code=400, detail="patches_json không phải JSON hợp lệ.")
    if not isinstance(patches, list) or not patches:
        raise HTTPException(status_code=400, detail="Danh sách patch rỗng.")

    ok_count = 0
    fails: list[dict] = []
    async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
        for p in patches[:500]:
            if not isinstance(p, dict):
                continue
            pid = p.get("id")
            point = p.get("point")
            note = str(p.get("note") or "").strip()
            link = str(p.get("link") or "").strip()
            if link:
                try:
                    link_norm = normalize_github_repo_url(link)
                    if link_norm:
                        link = link_norm
                except Exception:
                    pass
            try:
                pid_int = int(pid)
            except Exception:
                fails.append({"id": pid, "error": "id không hợp lệ"})
                continue
            try:
                point_f = float(point)
            except Exception:
                fails.append({"id": pid_int, "error": "point không hợp lệ"})
                continue
            point_f = max(0.0, min(100.0, point_f))
            point_str = _result_test_point_str(point_f)
            note_full = _normalize_result_test_note(note) if note else ""

            # Thử vài biến thể body: portal dùng point string; 500 hay gặp khi note quá dài hoặc link không khớp chuẩn hóa của Rikkei.
            attempt_bodies: list[dict] = []
            b_full: dict = {"point": point_str}
            if note_full:
                b_full["note"] = note_full
            if link:
                b_full["link"] = link
            attempt_bodies.append(b_full)
            if link:
                b_no_link = {"point": point_str}
                if note_full:
                    b_no_link["note"] = note_full
                attempt_bodies.append(b_no_link)
            attempt_bodies.append({"point": point_str})

            last_r: httpx.Response | None = None
            last_err: str = ""
            ok_this = False
            seen_keys: set[str] = set()
            for body in attempt_bodies:
                key = json.dumps(body, sort_keys=True, ensure_ascii=False)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                try:
                    r = await _rikkei_patch_result_test(client, tok, pid_int, body)
                    last_r = r
                    if r.status_code in (401, 403):
                        last_err = "Token không hợp lệ hoặc không có quyền"
                        break
                    if r.is_success:
                        ok_count += 1
                        ok_this = True
                        break
                    last_err = f"HTTP {r.status_code}"
                    snippet = (r.text or "")[:400].replace("\n", " ")
                    if snippet:
                        last_err += f" | {snippet}"
                except Exception as e:
                    last_err = str(e)[:400]

            if not ok_this:
                fail: dict = {"id": pid_int, "error": last_err or "PATCH thất bại"}
                if last_r is not None:
                    fail["rikkei_status"] = last_r.status_code
                    tb = (last_r.text or "").strip()
                    if tb:
                        fail["rikkei_body"] = tb[:500]
                fails.append(fail)

    return JSONResponse({"ok": True, "ok_count": ok_count, "fail_count": len(fails), "fails": fails[:50]})


@app.post("/api/hackathon-grade/export-xlsx")
async def api_hackathon_grade_export_xlsx(
    rows_json: str = Form(...),
) -> FileResponse:
    try:
        rows = json.loads(rows_json or "[]")
    except Exception:
        raise HTTPException(status_code=400, detail="rows_json không phải JSON hợp lệ.")
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows_json phải là list.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Hackathon"
    headers = ["Mã SV", "Họ tên", "Repo", "Link đề", "OK", "Điểm", "Nhận xét", "Lỗi repo", "Lỗi AI"]
    ws.append(headers)
    for r in rows[:5000]:
        if not isinstance(r, dict):
            continue
        ws.append(
            [
                str(r.get("studentCode") or ""),
                str(r.get("fullName") or ""),
                str(r.get("repo") or ""),
                str(r.get("assignment") or ""),
                "OK" if r.get("ok") else "FAIL",
                str(r.get("score") or ""),
                str(r.get("comment") or ""),
                str(r.get("repo_error") or ""),
                str(r.get("ai_error") or ""),
            ]
        )

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 28 if col in (3, 4, 7) else 16

    out_dir = tempfile.mkdtemp(prefix="hg_xlsx_")
    out_path = Path(out_dir) / "hackathon_results.xlsx"
    wb.save(str(out_path))
    background_tasks = BackgroundTasks()
    background_tasks.add_task(shutil.rmtree, out_dir, True)
    return FileResponse(
        path=str(out_path),
        filename=out_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=background_tasks,
    )


@app.post("/api/btvn")
async def api_btvn(
    background_tasks: BackgroundTasks,
    assignment_text: str = Form(""),
    assignment_image_urls: str = Form(""),
    submissions_text: str = Form(...),
    model: str = Form(""),
    github_token: str = Form(""),
    assignment_images: list[UploadFile] | None = File(None),
) -> JSONResponse:
    subs_lines = [
        ln.strip()
        for ln in (submissions_text or "").splitlines()
        if ln.strip() and not ln.strip().startswith("#")
    ]
    if not subs_lines:
        raise HTTPException(status_code=400, detail="Chưa có danh sách bài nộp (mỗi dòng 1 link GitHub).")

    imgs: list[tuple[str, bytes]] = []
    if assignment_images:
        for f in assignment_images:
            if not f or not f.filename:
                continue
            raw = await f.read()
            if not raw:
                continue
            ct = (f.content_type or "").strip() or "image/png"
            if not ct.startswith("image/"):
                continue
            # giới hạn nhẹ để tránh payload quá lớn
            if len(raw) > 5_000_000:
                raise HTTPException(status_code=400, detail=f"Ảnh quá lớn: {f.filename} (>5MB).")
            imgs.append((ct, raw))

    # Ảnh từ URL trong đề (từ Rikkei)
    ublob = (assignment_image_urls or "").strip()
    if ublob:
        try:
            urls = json.loads(ublob)
        except json.JSONDecodeError:
            urls = []
        if isinstance(urls, list):
            # giới hạn nhẹ để tránh payload
            for u in urls[:10]:
                if not isinstance(u, str):
                    continue
                got = await _rk_fetch_image_bytes(u)
                if got:
                    imgs.append(got)

    btvn_model = (model or "").strip() or (
        os.getenv("OPENROUTER_BTVN_MODEL", DEFAULT_BTVN_MODEL).strip()
        or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6").strip()
    )
    params = BtvnCommentParams(
        assignment_text=(assignment_text or "").strip(),
        assignment_images=imgs,
        submissions=subs_lines,
        model=btvn_model,
        github_token=(github_token or "").strip(),
    )

    loop = asyncio.get_event_loop()

    def work():
        return run_btvn_comments_json(params)

    ok, msg, rows = await loop.run_in_executor(_executor, work)
    if not ok or rows is None:
        raise HTTPException(status_code=500, detail=msg or "Lỗi không xác định.")

    at = (assignment_text or "").strip()
    fp = hashlib.sha1(at.encode("utf-8", errors="ignore")).hexdigest()[:10] if at else ""
    return JSONResponse(
        {
            "ok": True,
            "rows": rows,
            "assignment_fingerprint": {
                "chars": len(at),
                "sha1_10": fp,
                "head": at[:160],
            },
        }
    )


@app.post("/api/btvn/rikkei")
async def api_btvn_rikkei(
    rikkei_token: str = Form(...),
    class_id: str = Form(...),
    session_id: str = Form(...),
    course_id: str = Form(...),
    homework_id: str = Form(""),
    students_ids_json: str = Form("[]"),
    assignment_text: str = Form(""),
    assignment_image_urls: str = Form(""),
    model: str = Form(""),
    github_token: str = Form(""),
) -> JSONResponse:
    tok = (rikkei_token or "").strip()
    if not tok:
        raise HTTPException(status_code=400, detail="Thiếu token Rikkei.")
    cid = (class_id or "").strip()
    sid = (session_id or "").strip()
    crid = (course_id or "").strip()
    if not cid or not sid or not crid:
        raise HTTPException(status_code=400, detail="Thiếu class_id/session_id/course_id.")
    hid_req_raw = (homework_id or "").strip()
    hid_req: int | None = None
    if hid_req_raw:
        try:
            hid_req = int(hid_req_raw)
        except Exception:
            hid_req = None

    at = (assignment_text or "").strip()
    if not at:
        raise HTTPException(status_code=400, detail="Thiếu assignment_text (đề bài).")

    try:
        s_ids = json.loads(students_ids_json or "[]")
    except Exception:
        s_ids = []
    if not isinstance(s_ids, list):
        s_ids = []
    s_id_ints: list[int] = []
    for x in s_ids:
        try:
            ix = int(x)
        except Exception:
            continue
        s_id_ints.append(ix)

    # tải ảnh từ assignment_image_urls (nếu có)
    imgs: list[tuple[str, bytes]] = []
    ublob = (assignment_image_urls or "").strip()
    if ublob:
        try:
            urls = json.loads(ublob)
        except json.JSONDecodeError:
            urls = []
        if isinstance(urls, list):
            for u in urls[:10]:
                if not isinstance(u, str):
                    continue
                got = await _rk_fetch_image_bytes(u)
                if got:
                    imgs.append(got)

    btvn_model = (model or "").strip() or (
        os.getenv("OPENROUTER_BTVN_MODEL", DEFAULT_BTVN_MODEL).strip()
        or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6").strip()
    )

    try:
        students = _btvn_fetch_students(tok, cid, sid)
    except httpx.HTTPStatusError as e:
        raise HTTPException(
            status_code=502,
            detail=f"Lỗi API danh sách học sinh: {e.response.status_code} {e.response.text[:500]}",
        )
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Lỗi khi lấy danh sách học sinh: {e}")

    if s_id_ints:
        want = set(s_id_ints)
        students = [x for x in students if int(x.get("id") or -1) in want]

    target = []
    for st in students:
        sid_int = st.get("id")
        try:
            sid_int = int(sid_int)
        except Exception:
            continue
        target.append(
            {
                "studentId": sid_int,
                "studentCode": str(st.get("studentCode") or st.get("student_code") or "").strip(),
                "fullName": str(st.get("fullName") or st.get("full_name") or "").strip(),
            }
        )

    if not target:
        raise HTTPException(status_code=400, detail="Không có học sinh phù hợp để chấm.")

    # Build submissions for valid GitHub links only, aligned with target index.
    submission_links: list[str] = []
    submission_to_target_index: list[int] = []
    manual_rows: dict[int, dict] = {}

    for i, st in enumerate(target):
        st_student_id = st["studentId"]
        try:
            exercises = _btvn_fetch_all_exercises_for_student(
                tok,
                class_id=cid,
                course_id=crid,
                session_id=sid,
                student_id=st_student_id,
            )
        except Exception as e:
            manual_rows[i] = {
                "studentCode": st.get("studentCode") or "",
                "fullName": st.get("fullName") or "",
                "repo": "",
                "repo_error": f"Lỗi tải bài tập: {str(e)[:80]}",
                "comment": "",
                "ai_suspected": "Không",
                "ai_error": "",
            }
            continue

        chosen_link = ""
        if isinstance(exercises, list):
            for ex in exercises:
                hid = _btvn_homework_id(ex)
                if hid_req is not None and hid != hid_req:
                    continue
                link = ex.get("link_git") or ex.get("linkGit") or ""
                if isinstance(link, str):
                    link = link.strip()
                else:
                    link = ""
                if link:
                    chosen_link = link
                    break
        if not chosen_link:
            manual_rows[i] = {
                "studentCode": st.get("studentCode") or "",
                "fullName": st.get("fullName") or "",
                "repo": "",
                "repo_error": "Chưa có link Git.",
                "comment": "",
                "ai_suspected": "Không",
                "ai_error": "",
            }
        else:
            submission_to_target_index.append(i)
            submission_links.append(chosen_link)

    rows_out: list[dict] = []
    # init with manual rows
    rows_out = [manual_rows.get(i) or None for i in range(len(target))]

    if submission_links:
        params = BtvnCommentParams(
            assignment_text=at,
            assignment_images=imgs,
            submissions=submission_links,
            model=btvn_model,
            github_token=(github_token or "").strip(),
        )
        loop = asyncio.get_event_loop()

        def work():
            return run_btvn_comments_json(params)

        ok, msg, rows = await loop.run_in_executor(_executor, work)
        if not ok or rows is None:
            raise HTTPException(status_code=500, detail=msg or "Lỗi chấm BTVN")

        for j, r in enumerate(rows):
            ti = submission_to_target_index[j] if j < len(submission_to_target_index) else None
            if ti is None:
                continue
            row = dict(r or {})
            row["studentCode"] = target[ti].get("studentCode") or ""
            row["fullName"] = target[ti].get("fullName") or ""
            rows_out[ti] = row

    default_row = {
        "studentCode": "",
        "fullName": "",
        "repo": "",
        "repo_error": "Không xác định được kết quả.",
        "comment": "",
        "ai_suspected": "Không",
        "ai_error": "",
    }
    rows_out_final: list[dict] = []
    for i in range(len(target)):
        v = rows_out[i] if i < len(rows_out) else None
        if isinstance(v, dict):
            rows_out_final.append(v)
        else:
            rows_out_final.append(
                {**default_row, "studentCode": target[i].get("studentCode") or "", "fullName": target[i].get("fullName") or ""}
            )

    at_fp = (assignment_text or "").strip()
    fp = hashlib.sha1(at_fp.encode("utf-8", errors="ignore")).hexdigest()[:10] if at_fp else ""
    return JSONResponse(
        {
            "ok": True,
            "rows": rows_out_final,
            "assignment_fingerprint": {
                "chars": len(at_fp),
                "sha1_10": fp,
                "head": at_fp[:160],
            },
        }
    )


@app.post("/api/group-activity")
async def api_group_activity(
    video_transcript: str = Form(...),
    report_file: UploadFile = File(...),
    model: str = Form(""),
) -> PlainTextResponse:
    try:
        from cham_bai.settings import api_key as _need_key

        _need_key()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))

    m = (model or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6")).strip()
    vt = (video_transcript or "").strip()
    if not vt:
        raise HTTPException(status_code=400, detail="Thiếu transcript/ghi chú video.")

    if not report_file or not report_file.filename:
        raise HTTPException(status_code=400, detail="Thiếu file báo cáo.")
    raw = await report_file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="File báo cáo rỗng.")
    if len(raw) > 12_000_000:
        raise HTTPException(status_code=400, detail="File báo cáo quá lớn (>12MB).")
    fname = str(report_file.filename or "").strip()
    params = GroupGradeParams(
        report_filename=fname,
        report_bytes=raw,
        video_transcript=vt,
        model=m,
    )

    loop = asyncio.get_event_loop()

    def work():
        return grade_group_activity(params)

    try:
        result = await loop.run_in_executor(_executor, work)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return PlainTextResponse(result or "", media_type="text/plain; charset=utf-8")


@app.post("/api/reading")
async def api_reading(
    background_tasks: BackgroundTasks,
    subject: str = Form(...),
    session: str = Form(...),
    lesson: str = Form(...),
    session_stt: str = Form("1"),
    lesson_stt: str = Form("1"),
    video_url: str = Form(""),
    technology: str = Form(""),
    audience: str = Form("Sinh viên năm 1, mới học lập trình"),
    learning_goals: str = Form(DEFAULT_LEARNING_GOALS),
    references_hint: str = Form(""),
    text_model: str = Form(""),
    image_model: str = Form(""),
    generate_illustrations: str = Form("true"),
) -> FileResponse:
    subject = subject.strip()
    session = session.strip()
    lesson = lesson.strip()
    if not subject or not session or not lesson:
        raise HTTPException(status_code=400, detail="Điền đủ môn, session và lesson.")

    try:
        from cham_bai.settings import api_key as _need_key

        _need_key()
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e

    ss = (session_stt or "1").strip() or "1"
    ls = (lesson_stt or "1").strip() or "1"
    stem = reading_output_stem(subject, ss, ls)

    work_dir = tempfile.mkdtemp(prefix="reading_")
    out_docx = Path(work_dir) / f"{stem}.docx"
    out_xlsx = Path(work_dir) / f"{stem}.xlsx"

    params = ReadingDocParams(
        subject=subject,
        session=session,
        lesson=lesson,
        session_stt=ss,
        lesson_stt=ls,
        video_url=(video_url or "").strip() or None,
        technology=(technology or "").strip(),
        audience=(audience or "").strip(),
        learning_goals=(learning_goals or "").strip(),
        references_hint=(references_hint or "").strip(),
        text_model=(text_model or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6")).strip(),
        image_model=(image_model or IMAGE_MODEL_OPTIONS[0]).strip(),
        generate_illustrations=_parse_bool_form(generate_illustrations),
        output_docx=out_docx,
        output_xlsx=out_xlsx,
    )

    loop = asyncio.get_event_loop()

    def gen():
        return run_reading_generation(params)

    ok, msg = await loop.run_in_executor(_executor, gen)
    if not ok:
        shutil.rmtree(work_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=msg)

    zip_fd, zip_path = tempfile.mkstemp(suffix=".zip")
    os.close(zip_fd)
    zip_p = Path(zip_path)
    with zipfile.ZipFile(zip_p, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(out_docx, arcname=out_docx.name)
        zf.write(out_xlsx, arcname=out_xlsx.name)

    def cleanup():
        try:
            zip_p.unlink(missing_ok=True)
        except OSError:
            pass
        shutil.rmtree(work_dir, ignore_errors=True)

    background_tasks.add_task(cleanup)

    zip_name = sanitize_reading_filename_part(stem) + ".zip"
    return FileResponse(
        path=str(zip_p),
        filename=zip_name,
        media_type="application/zip",
    )


@app.post("/api/hackathon")
async def api_hackathon(
    duration_minutes: int = Form(120),
    mode: str = Form("manual"),  # manual | ai
    subject: str = Form(""),
    outline_text: str = Form(""),
    technology: str = Form("MySQL"),
    ide: str = Form("MySQL Workbench"),
    exam_code: str = Form("006"),
    exam_template: str = Form("mysql"),
    extra_notes: str = Form(""),
    model: str = Form(""),
    body_text: str = Form(""),
) -> FileResponse:
    mins = max(1, int(duration_minutes or 0))
    m = (model or os.getenv("OPENROUTER_MODEL", "anthropic/claude-sonnet-4.6")).strip()

    raw: bytes
    if (mode or "").strip().lower() == "ai":
        # Need OpenRouter key
        try:
            from cham_bai.settings import api_key as _need_key

            _need_key()
        except RuntimeError as e:
            raise HTTPException(status_code=400, detail=str(e))

        # Fixed header like mẫu
        header_top = "KIỂM TRA HACKATHON"
        subj = (subject or "").strip() or "NHẬP MÔN CSDL MYSQL"
        try:
            ex_int = int(str(exam_code or "").strip() or "6")
        except Exception:
            ex_int = 6
        ex_int = max(1, min(999, ex_int))
        ex_code = f"{ex_int:03d}"
        hs = f"{subj} - Đề {ex_code}"
        tpl = normalize_exam_template(exam_template)

        try:
            spec = generate_hackathon_exam_spec(
                model=m,
                header_top=header_top,
                header_sub=hs,
                duration_minutes=mins,
                subject=subj,
                outline_text=(outline_text or "").strip(),
                technology=(technology or "").strip() or "MySQL",
                ide=(ide or "").strip() or "MySQL Workbench",
                exam_code=ex_code,
                exam_template=tpl,
                extra_notes=(extra_notes or "").strip(),
            )
            # Ensure required "Yêu cầu" always exists (AI sometimes omits).
            req_bullets = build_required_bullets(
                exam_template=tpl,
                technology=(technology or "").strip() or "MySQL",
                ide=(ide or "").strip() or "MySQL Workbench",
            )
            spec = ensure_required_section(spec, required_bullets=req_bullets)
            spec = ensure_practice_section(spec)
            raw = build_hackathon_exam_docx_from_spec(spec)
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Lỗi tạo đề bằng AI: {e}")
    else:
        body = (body_text or "").strip()
        if not body:
            raise HTTPException(status_code=400, detail="Thiếu nội dung đề.")
        header_top = "KIỂM TRA HACKATHON"
        subj = (subject or "").strip() or "NHẬP MÔN CSDL MYSQL"
        try:
            ex_int = int(str(exam_code or "").strip() or "6")
        except Exception:
            ex_int = 6
        ex_int = max(1, min(999, ex_int))
        ex_code = f"{ex_int:03d}"
        hs = f"{subj} - Đề {ex_code}"
        params = HackathonExamParams(
            header_top=header_top,
            header_sub=hs,
            duration_minutes=mins,
            body_text=body,
        )
        raw = build_hackathon_exam_docx_bytes(params)

    if not raw:
        raise HTTPException(status_code=500, detail="Không tạo được file DOCX.")

    out_dir = tempfile.mkdtemp(prefix="hackathon_out_")
    out_path = Path(out_dir) / "de_hackathon.docx"
    out_path.write_bytes(raw)
    background_tasks = BackgroundTasks()
    background_tasks.add_task(shutil.rmtree, out_dir, True)
    return FileResponse(
        path=str(out_path),
        filename=out_path.name,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        background=background_tasks,
    )


def main() -> None:
    import uvicorn

    uvicorn.run(
        "cham_bai.web_app:app",
        host="127.0.0.1",
        port=8765,
        reload=False,
    )


if __name__ == "__main__":
    main()
