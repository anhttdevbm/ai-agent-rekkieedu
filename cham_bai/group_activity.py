from __future__ import annotations

import json
import re
from dataclasses import dataclass
from typing import Any

import io
import html as _html
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from cham_bai.openrouter import complete_chat_raw


@dataclass
class GroupGradeParams:
    report_filename: str
    report_bytes: bytes
    video_transcript: str
    model: str


@dataclass
class GroupMemberRow:
    name: str
    attendance: str  # present | absent | unknown
    note: str = ""


@dataclass
class GroupGradeResult:
    comment: str
    members: list[GroupMemberRow]


_SYSTEM = """Bạn là giảng viên chấm HOẠT ĐỘNG NHÓM.

Đầu vào: BÁO CÁO (text trích) + GHI CHÚ NỘI DUNG VIDEO (phụ đề/ghi chú do người chấm cung cấp).

Nguyên tắc chấm (nội bộ, không lặp lại trong nhận xét):
- Không bịa nội dung video; thiếu dữ liệu thì nói thẳng «video/ghi chú chưa đủ».
- Đối chiếu báo cáo với video: phối hợp, phân công, nhóm trưởng điều phối; báo cáo có đủ công việc đã làm, khó khăn, thành viên vắng/chưa làm.

YÊU CẦU XUẤT (bắt buộc):
- Chỉ 1 đoạn: tối đa 2 câu, khoảng 30–55 từ tiếng Việt (ngắn hơn nữa nếu đủ ý). Không JSON, bullet, markdown.
- Giọng thẳng: gọi đúng tên (vd. Hoàng, Lê Tuấn Đạo); cấm «một thành viên», «có thể», «dường như».
- Chỉ dùng «video» (cấm transcript/phụ đề). Không mở đầu dài «Nhóm trưởng X báo cáo…» — vào thẳng kết luận.
- Câu 1 (bắt buộc): điểm đạt + thiếu sót chính trong cùng một câu (dùng dấu phẩy/chấm phẩy, không tách thành nhiều câu kể lể).
- Câu 2 (chỉ khi cần): một việc phải sửa, ≤12 từ. Nếu nhóm tốt: chỉ 1 câu khen ngắn.
- Không liệt kê từng bài/Session trừ khi thiếu sót gắn trực tiếp; ưu tiên 1–2 tên người quan trọng nhất."""

_MEMBERS_SYSTEM = """Bạn trích danh sách thành viên nhóm từ BÁO CÁO (chỉ đọc báo cáo, không dùng video/transcript).

Trả về DUY NHẤT một JSON object hợp lệ, không markdown:
{"members": [{"name": "Họ tên đầy đủ trong báo cáo", "attendance": "present"|"absent"|"unknown", "note": ""}]}

Quy ước attendance:
- present: báo cáo ghi có mặt, tham gia đủ, hoàn thành, có trong buổi họp
- absent: vắng họp, không tham gia, nghỉ, không hoàn thành (theo báo cáo)
- unknown: có tên nhưng báo cáo không nói rõ có mặt hay vắng

note: lý do vắng/ghi chú ngắn nếu báo cáo có, không thì "".
Không thêm người không xuất hiện trong báo cáo. Giữ đúng họ tên tiếng Việt."""


def _strip_xml_tags(s: str) -> str:
    # keep only text nodes content (very basic)
    s = re.sub(r"<[^>]+>", " ", s)
    s = _html.unescape(s)
    s = re.sub(r"[ \t\r\f\v]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def _wb_to_text(xlsx_bytes: bytes, *, max_cells: int = 2500) -> str:
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)  # type: ignore[name-defined]
    parts: list[str] = []
    cells = 0
    for ws in wb.worksheets[:8]:
        parts.append(f"=== SHEET: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            if cells >= max_cells:
                parts.append("[... TRUNCATED ...]")
                return "\n".join(parts).strip()
            vals = []
            for v in row:
                if v is None:
                    vals.append("")
                else:
                    s = str(v).strip()
                    vals.append(s)
            line = " | ".join(x for x in vals if x != "")
            if line.strip():
                parts.append(line)
            cells += 1
    return "\n".join(parts).strip()

def _extract_docx_plain_bytes(raw: bytes, *, max_chars: int = 28000) -> str:
    doc = Document(io.BytesIO(raw))
    parts: list[str] = []
    for p in doc.paragraphs:
        t = (p.text or "").strip()
        if t:
            parts.append(t)
    text = "\n".join(parts).strip()
    if len(text) > max_chars:
        text = text[: max_chars - 80] + " … [TRUNCATED]"
    return text


def report_file_bytes_to_text(filename: str, raw: bytes) -> tuple[str, list[str]]:
    warns: list[str] = []
    name = (filename or "").strip()
    ext = Path(name).suffix.lower()
    if ext == ".xlsx":
        try:
            return _wb_to_text(raw), warns
        except Exception as e:
            return "", [f"Lỗi đọc XLSX: {e}"]
    if ext == ".docx":
        try:
            return _extract_docx_plain_bytes(raw), warns
        except Exception as e:
            return "", [f"Lỗi đọc DOCX: {e}"]
    return "", [f"File báo cáo chưa hỗ trợ: {name or '(không tên)'} (chỉ nhận .docx hoặc .xlsx)"]


def _clean_plain_paragraph(s: str) -> str:
    t = (s or "").strip()
    # remove common markdown fences if provider adds them
    t = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", t)
    t = re.sub(r"\s*```$", "", t)
    # collapse whitespace
    t = re.sub(r"[ \t\r\f\v]+", " ", t)
    t = re.sub(r"\n{2,}", "\n", t)
    t = t.replace("\n", " ").strip()
    return t


_GROUP_COMMENT_MAX_WORDS = 55
_GROUP_COMMENT_MAX_SENTENCES = 2


def _truncate_words(s: str, max_words: int) -> str:
    words = s.split()
    if len(words) <= max_words:
        return s
    cut = " ".join(words[:max_words])
    for sep in (". ", "! ", "? ", "; ", "。"):
        pos = cut.rfind(sep)
        if pos > len(cut) * 0.35:
            return cut[: pos + 1].strip()
    return cut.rstrip(" ,;:") + "."


def _polish_group_comment(s: str) -> str:
    """Rút gọn và chuẩn từ ngữ nhận xét gửi sinh viên."""
    t = _clean_plain_paragraph(s)
    if not t:
        return t
    t = re.sub(r"\btranscript\b", "video", t, flags=re.IGNORECASE)
    t = re.sub(r"\bphụ\s*đề\b", "video", t, flags=re.IGNORECASE)
    t = re.sub(r"\btimedtext\b", "video", t, flags=re.IGNORECASE)
    t = re.sub(r"\bvideo\s+video\b", "video", t, flags=re.IGNORECASE)
    t = re.sub(r"[ \t\r\f\v]+", " ", t).strip()

    # Tối đa 2 câu.
    sentences = re.split(r"(?<=[.!?…])\s+", t)
    sentences = [x.strip() for x in sentences if x.strip()]
    if len(sentences) > _GROUP_COMMENT_MAX_SENTENCES:
        t = " ".join(sentences[:_GROUP_COMMENT_MAX_SENTENCES])
    else:
        t = " ".join(sentences)

    return _truncate_words(t, _GROUP_COMMENT_MAX_WORDS)


def _parse_json_object(raw: str) -> dict[str, Any] | None:
    t = (raw or "").strip()
    t = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", t)
    t = re.sub(r"\s*```$", "", t).strip()
    start = t.find("{")
    end = t.rfind("}")
    if start < 0 or end <= start:
        return None
    try:
        obj = json.loads(t[start : end + 1])
    except json.JSONDecodeError:
        return None
    return obj if isinstance(obj, dict) else None


def _normalize_attendance(v: Any) -> str:
    s = str(v or "").strip().lower()
    if s in ("present", "co_mat", "có mặt", "co mat", "đủ", "du", "yes", "true", "1"):
        return "present"
    if s in ("absent", "vang", "vắng", "khong", "không", "no", "false", "0", "nghi"):
        return "absent"
    return "unknown"


def extract_members_from_report(report_text: str, *, model: str) -> list[GroupMemberRow]:
    """Trích thành viên + có mặt/vắng chỉ từ text báo cáo."""
    rt = (report_text or "").strip()
    if not rt or rt.startswith("("):
        return []

    user = "BÁO CÁO (text trích):\n" + rt[:24000]
    text, _ = complete_chat_raw(
        [{"role": "system", "content": _MEMBERS_SYSTEM}, {"role": "user", "content": user}],
        model=model,
        temperature=0.1,
        max_tokens=900,
        timeout_s=180.0,
    )
    blob = _parse_json_object(text)
    if not blob:
        return []
    raw_list = blob.get("members")
    if not isinstance(raw_list, list):
        return []

    out: list[GroupMemberRow] = []
    seen: set[str] = set()
    for item in raw_list:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        if not name or len(name) < 2:
            continue
        key = name.casefold()
        if key in seen:
            continue
        seen.add(key)
        note = str(item.get("note") or "").strip()
        out.append(
            GroupMemberRow(
                name=name,
                attendance=_normalize_attendance(item.get("attendance")),
                note=note[:200],
            )
        )
    return out


def _grade_comment_only(
    *,
    report_text: str,
    video_notes: str,
    warns: list[str],
    model: str,
) -> str:
    user_parts = [
        {
            "type": "text",
            "text": (
                "BÁO CÁO (text trích):\n"
                + (report_text.strip() or "(trống/không đọc được)")
                + "\n\nGHI CHÚ NỘI DUNG VIDEO (phụ đề/ghi chú):\n"
                + (video_notes or "(trống)")
                + ("\n\nCẢNH BÁO:\n" + "\n".join(warns) if warns else "")
            ),
        }
    ]
    text, _ = complete_chat_raw(
        [{"role": "system", "content": _SYSTEM}, {"role": "user", "content": user_parts}],
        model=model,
        temperature=0.12,
        max_tokens=120,
        timeout_s=300.0,
    )
    out = _polish_group_comment(text)
    if not out:
        return "Không tạo được nhận xét (model trả rỗng)."
    return out


def grade_group_activity(params: GroupGradeParams) -> GroupGradeResult:
    report_text, warns = report_file_bytes_to_text(params.report_filename, params.report_bytes)
    video_notes = (params.video_transcript or "").strip()
    members = extract_members_from_report(report_text, model=params.model)
    comment = _grade_comment_only(
        report_text=report_text,
        video_notes=video_notes,
        warns=warns,
        model=params.model,
    )
    return GroupGradeResult(comment=comment, members=members)


def group_grade_result_to_dict(result: GroupGradeResult) -> dict[str, Any]:
    return {
        "comment": result.comment,
        "members": [
            {
                "name": m.name,
                "attendance": m.attendance,
                "note": m.note,
            }
            for m in result.members
        ],
    }

