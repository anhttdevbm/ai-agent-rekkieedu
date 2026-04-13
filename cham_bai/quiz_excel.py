from __future__ import annotations

import math
import re
import shutil
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Bộ 5 câu cố định: (STT, Mức độ — đúng tên hiển thị, Mục tiêu).
# Mức độ chỉ gồm: Thông hiểu (2 câu), Vận dụng sơ bộ, Phân biệt/So sánh, Phân tích sơ bộ — không dùng "Nhớ", "Hiểu", "Vận dụng" chung chung.
QUIZ_FIXED_LEVELS: list[tuple[int, str, str]] = [
    (1, "Thông hiểu", "Nhận diện cú pháp"),
    (2, "Thông hiểu", "Hiểu cơ chế hoạt động"),
    (3, "Vận dụng sơ bộ", "Đọc hiểu code"),
    (4, "Phân biệt/So sánh", "Tránh nhầm lẫn"),
    (5, "Phân tích sơ bộ", "Dự đoán kết quả có bẫy"),
]


def _norm_key(s: str) -> str:
    s = unicodedata.normalize("NFC", (s or "").strip().lower())
    s = re.sub(r"\s+", " ", s)
    return s


_EXPL_PREFIX_RES = (
    re.compile(r"^(đúng|sai|dung)\s*[:.;，,–\-]\s*", re.I),
    re.compile(r"^(đúng|sai|dung)\s+", re.I),
    re.compile(r"^(đáp án\s*)?(đúng|sai|dung)\s*[:.;，,–\-]?\s+", re.I),
)


def clean_quiz_explanation(text: str) -> str:
    """Giải thích chỉ giữ nội dung lý do — bỏ tiền tố Đúng/Sai (cột Kết quả đã thể hiện)."""
    s = (text or "").strip()
    for _ in range(4):
        hit = False
        for pat in _EXPL_PREFIX_RES:
            m = pat.match(s)
            if m:
                s = s[m.end() :].strip()
                hit = True
                break
        if not hit:
            break
    return s.strip()


_QUIZ_FONT_NAME = "Calibri"
_QUIZ_FONT_SIZE = 11


def _column_width_excel(ws: Any, col_idx: int, fallback: float) -> float:
    letter = get_column_letter(col_idx)
    dim = ws.column_dimensions.get(letter)
    w = dim.width if dim is not None and dim.width is not None else None
    return float(w) if w is not None else fallback


def _chars_per_line_from_width(width_excel: float) -> float:
    """Excel column width ≈ số ký tự Calibri 11 gần đúng; hệ số <1 vì chữ Việt/rộng."""
    # Hệ số thấp hơn → ước lượng nhiều dòng wrap hơn → cao hàng an toàn hơn (tránh cắt chữ)
    return max(6.0, float(width_excel) * 0.82)


def _estimate_wrapped_line_count(text: str, col_width_excel: float) -> int:
    """Số dòng hiển thị khi wrap trong cột có độ rộng `col_width_excel` (không chỉ \\n)."""
    if not (text or "").strip():
        return 1
    cpl = _chars_per_line_from_width(col_width_excel)
    total = 0
    for line in str(text).splitlines():
        line = line.strip()
        if not line:
            total += 1
            continue
        total += max(1, math.ceil(len(line) / cpl))
    return max(1, total)


def _cell_plain_for_width(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, CellRichText):
        parts: list[str] = []
        for p in val:
            if isinstance(p, str):
                parts.append(p)
            elif isinstance(p, TextBlock):
                parts.append(p.text or "")
            else:
                parts.append(str(p))
        return "".join(parts)
    return str(val)


def _apply_ket_qua_cell(cell: Any, kq: str) -> None:
    """Đúng/Sai cùng font Calibri; chỉ «Đúng» in đậm (Rich Text). «Sai» b=False để Excel không herit đậm."""
    t = (kq or "").strip()
    nk = _norm_key(t)
    ok = nk in ("đúng", "dung", "true", "yes")
    fn, fs = _QUIZ_FONT_NAME, float(_QUIZ_FONT_SIZE)
    normal = InlineFont(rFont=fn, sz=fs, b=False)
    if ok:
        cell.value = CellRichText(TextBlock(InlineFont(rFont=fn, sz=fs, b=True), "Đúng"))
    else:
        cell.value = CellRichText(TextBlock(normal, "Sai"))
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.font = Font(name=fn, size=_QUIZ_FONT_SIZE, bold=False)


def autofit_quiz_columns_and_rows(
    ws,
    *,
    min_col_width: float = 12.0,
    max_col_width: float = 92.0,
    header_row: int = 1,
    vertical_blocks: list[VerticalQuizBlock] | None = None,
    c_q: int | None = None,
    c_ans: int | None = None,
    c_res: int | None = None,
    c_exp: int | None = None,
) -> None:
    """Ước lượng độ rộng cột + chiều cao hàng (kể cả wrap mềm) để nội dung không bị cắt."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    for col_idx in range(1, max_col + 1):
        best = min_col_width
        floor = min_col_width
        if c_exp is not None and col_idx == c_exp:
            floor = max(floor, 38.0)
        for r in range(1, max_row + 1):
            cell = ws.cell(r, col_idx)
            if isinstance(cell, MergedCell):
                continue
            s = _cell_plain_for_width(cell.value)
            if not s:
                continue
            longest_line = max((len(line) for line in s.splitlines()), default=0)
            est = min(max(longest_line * 1.08 + 3.0, floor, min_col_width), max_col_width)
            best = max(best, est)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(best, floor)

    line_height_pt = 15.5

    for r in range(1, max_row + 1):
        max_lines = 1
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            s = _cell_plain_for_width(cell.value)
            if not s:
                continue
            wcol = _column_width_excel(ws, c, min_col_width)
            max_lines = max(max_lines, _estimate_wrapped_line_count(s, wcol))
        if r == header_row:
            max_lines = max(max_lines, 2)
        pts = line_height_pt * max_lines + 14.0
        ws.row_dimensions[r].height = min(max(pts, 18.0), 409.0)

    if vertical_blocks and c_q and c_ans and c_res is not None and c_exp:
        w_q = _column_width_excel(ws, c_q, min_col_width)
        w_a = _column_width_excel(ws, c_ans, min_col_width)
        w_e = _column_width_excel(ws, c_exp, min_col_width)
        labels = ("A", "B", "C", "D")
        for bi, blk in enumerate(vertical_blocks):
            r0 = 2 + bi * 4
            q_lines = _estimate_wrapped_line_count(str(blk.question_text), w_q)
            q_share = max(1, math.ceil(q_lines / 4))
            for j in range(4):
                r = r0 + j
                text, _kq, expl = blk.choices[j]
                prefix = f"{labels[j]}. " if not str(text).lstrip().startswith(f"{labels[j]}.") else ""
                ans_s = f"{prefix}{text}".strip()
                le = _estimate_wrapped_line_count(clean_quiz_explanation(expl), w_e)
                la = _estimate_wrapped_line_count(ans_s, w_a)
                res_cell = ws.cell(r, c_res)
                lr = _estimate_wrapped_line_count(
                    _cell_plain_for_width(res_cell.value),
                    _column_width_excel(ws, c_res, 12.0),
                )
                row_lines = max(le, la, lr, q_share, 1)
                need = line_height_pt * row_lines + 18.0
                cur = ws.row_dimensions[r].height
                cur_f = float(cur) if cur is not None else 0.0
                ws.row_dimensions[r].height = min(max(need, cur_f), 409.0)


def _header_norms(headers: list[str]) -> list[str]:
    return [_norm_key(h) for h in headers]


def _norms_contain_any(norms: list[str], *candidates: str) -> bool:
    return any(c in norms for c in candidates)


def is_vertical_quiz_template(headers: list[str]) -> bool:
    """Mẫu 7 cột quiz dọc. Cho phép cột A là «Câu hỏi» (số câu) và cột D cũng «Câu hỏi» (nội dung)."""
    if len(headers) < 7:
        return False
    norms = _header_norms(headers)
    if "mức độ" not in norms or "mục tiêu" not in norms:
        return False
    if not _norms_contain_any(norms, "các đáp án", "các đáp an", "đáp án"):
        return False
    if not _norms_contain_any(norms, "kết quả", "ket qua"):
        return False
    if not _norms_contain_any(
        norms,
        "giải thích",
        "giai thích",
        "giảithích",
        "giai thich",
        "ghi chú",
        "ghi chu",
    ):
        return False
    n_ch = sum(1 for n in norms if n == "câu hỏi")
    has_stt = "stt" in norms or "số thứ tự" in norms or "số tt" in norms
    if n_ch >= 2:
        return True
    if n_ch == 1 and has_stt:
        return True
    return False


def resolve_vertical_column_indexes(headers: list[str]) -> tuple[int, int, int, int, int, int, int]:
    """
    Trả về chỉ số cột 1-based: (stt_số_câu, mức_độ, mục_tiêu, nội_dung_câu_hỏi, đáp_án, kết_quả, giải_thích).
    Xử lý mẫu có hai cột cùng chữ «Câu hỏi»: cột đầu = STT/số câu, cột sau (cùng tên) = nội dung.
    """
    if len(headers) < 7:
        raise ValueError("Mẫu quiz dọc cần ít nhất 7 cột tiêu đề.")
    norms = _header_norms(headers)

    def idx_of(norm: str) -> int:
        try:
            return norms.index(norm) + 1
        except ValueError as e:
            raise ValueError(f"Thiếu tiêu đề cột «{norm}» trong file mẫu.") from e

    def idx_of_any(*candidates: str) -> int:
        for cand in candidates:
            try:
                return norms.index(cand) + 1
            except ValueError:
                continue
        raise ValueError(
            f"Thiếu cột (thử các tên: {', '.join(candidates)}) trong hàng tiêu đề file mẫu."
        )

    c_md = idx_of("mức độ")
    c_mt = idx_of("mục tiêu")
    c_ans = idx_of_any("các đáp án", "các đáp an", "đáp án", "dap an")
    c_res = idx_of_any("kết quả", "ket qua")
    c_exp = idx_of_any(
        "giải thích",
        "giai thích",
        "giảithích",
        "giai thich",
        "giai thich cho lua chon",
        "ghi chú",
        "ghi chu",
    )

    ch_positions = [i for i, n in enumerate(norms) if n == "câu hỏi"]
    if len(ch_positions) >= 2:
        c_stt = ch_positions[0] + 1
        c_q = ch_positions[1] + 1
    elif "stt" in norms:
        c_stt = idx_of("stt")
        if not ch_positions:
            raise ValueError("Mẫu có STT nhưng thiếu cột «Câu hỏi» (nội dung).")
        c_q = ch_positions[0] + 1
    elif "số thứ tự" in norms:
        c_stt = idx_of("số thứ tự")
        if not ch_positions:
            raise ValueError("Thiếu cột «Câu hỏi» nội dung.")
        c_q = ch_positions[0] + 1
    elif len(ch_positions) == 1:
        raise ValueError(
            "File mẫu chỉ có một cột «Câu hỏi». Cần thêm cột STT hoặc thêm cột «Câu hỏi» thứ hai "
            "cho nội dung (như mẫu: A= số câu, D= nội dung)."
        )
    else:
        raise ValueError("Không xác định được cột số câu / nội dung câu hỏi.")

    return (c_stt, c_md, c_mt, c_q, c_ans, c_res, c_exp)


@dataclass(frozen=True)
class VerticalQuizBlock:
    """Một câu quiz dọc; choices = 4 bộ (nội dung đáp án, 'Đúng'|'Sai', giải thích)."""

    stt: int
    muc_do: str
    muc_tieu: str
    question_text: str
    choices: tuple[tuple[str, str, str], tuple[str, str, str], tuple[str, str, str], tuple[str, str, str]]


def default_template_path() -> Path:
    """Đường dẫn file mẫu đi kèm package."""
    return Path(__file__).resolve().parent / "templates" / "quiz_mau.xlsx"


def _package_repo_root() -> Path:
    """Thư mục gốc dự án (chứa `cham_bai/` và `example/`)."""
    return Path(__file__).resolve().parent.parent


def lesson_quiz_example_template_path() -> Path:
    """Mẫu Excel cho quiz theo lesson: `example/quizz-lession-example.xlsx`."""
    return _package_repo_root() / "example" / "quizz-lession-example.xlsx"


def session_warmup_quiz_example_template_path() -> Path:
    """Mẫu Excel cho quiz session đầu giờ: `example/Quizz_Session_Dau_Gio_Example.xlsx`."""
    return _package_repo_root() / "example" / "Quizz_Session_Dau_Gio_Example.xlsx"


def ensure_session_warmup_quiz_example_template() -> Path:
    """
    Đảm bảo có `example/Quizz_Session_Dau_Gio_Example.xlsx`.
    Nếu chưa có, tạo mới đúng header form warmup (1 hàng / câu).
    """
    p = session_warmup_quiz_example_template_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.is_file():
        wb = Workbook()
        ws = wb.active
        ws.title = "Quiz"
        # Form warmup (1 hàng / câu)
        ws.append(
            [
                "question_content",
                "answer_1",
                "explanation_answer_1",
                "answer_2",
                "explanation_answer_2",
                "answer_3",
                "explanation_answer_3",
                "answer_4",
                "explanation_answer_4",
                "isCorrect",
                "difficulty",
                "category",
            ]
        )
        wb.save(p)
    return p


def fill_template_session_warmup_quiz(
    template_xlsx: Path,
    output_xlsx: Path,
    rows: list[dict[str, object]],
) -> None:
    """
    Ghi quiz warmup (45 câu) vào file output theo template:
    question_content, answer_1..4 + explanation_answer_1..4, isCorrect (1..4), difficulty (int), category.
    30 câu đầu: BÀI CŨ; 15 câu sau: BÀI MỚI.
    """
    wb = load_workbook(template_xlsx)
    ws = wb.active
    ws.cell(1, 12).value = "category"
    # Ghi đè từ hàng 2
    start_row = 2
    for i, r in enumerate(rows):
        rr = start_row + i
        ws.cell(rr, 1).value = r.get("question_content", "")
        ws.cell(rr, 2).value = r.get("answer_1", "")
        ws.cell(rr, 3).value = r.get("explanation_answer_1", "")
        ws.cell(rr, 4).value = r.get("answer_2", "")
        ws.cell(rr, 5).value = r.get("explanation_answer_2", "")
        ws.cell(rr, 6).value = r.get("answer_3", "")
        ws.cell(rr, 7).value = r.get("explanation_answer_3", "")
        ws.cell(rr, 8).value = r.get("answer_4", "")
        ws.cell(rr, 9).value = r.get("explanation_answer_4", "")
        ws.cell(rr, 10).value = r.get("isCorrect", "")
        ws.cell(rr, 11).value = r.get("difficulty", "")
        ws.cell(rr, 12).value = "BÀI CŨ" if i < 30 else "BÀI MỚI"
    wb.save(output_xlsx)


def ensure_lesson_quiz_example_template() -> Path:
    """
    Đảm bảo có `example/quizz-lession-example.xlsx`.
    Nếu chưa có, sao chép từ mẫu chuẩn (cùng khung 7 cột dọc).
    """
    p = lesson_quiz_example_template_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    if not p.is_file():
        shutil.copy2(ensure_default_quiz_template(), p)
    return p


def ensure_default_quiz_template() -> Path:
    """Tạo hoặc nâng cấp file mẫu mặc định (7 cột khung dọc) nếu chưa có hoặc đang là bản cũ."""
    p = default_template_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    if p.is_file():
        try:
            if is_vertical_quiz_template(read_headers_from_template(p)):
                return p
        except Exception:
            pass
    wb = Workbook()
    ws = wb.active
    ws.title = "Quiz"
    # Giống mẫu giảng viên: A–D merge meta + nội dung câu; A và D cùng nhãn «Câu hỏi» (số vs nội dung).
    headers = [
        "Câu hỏi",
        "Mức độ",
        "Mục tiêu",
        "Câu hỏi",
        "Các đáp án",
        "Kết quả",
        "Giải thích",
    ]
    ws.append(headers)
    wb.save(p)
    return p


def read_header_row(ws, row: int = 1) -> list[str]:
    """
    Đọc tiêu đề hàng `row` (trái → phải), bỏ qua ô trống nhưng không dừng sớm vì 1–3 cột trống giữa các nhãn
    (trường hợp hay gặp: còn vài cột trống trước «Giải thích» → bản cũ đọc thiếu cột).
    Dừng khi đã có ít nhất một tiêu đề và gặp ≥20 ô trống liên tiếp, hoặc hết dải quét.
    """
    headers: list[str] = []
    empty_run = 0
    last_non_empty_col = 0
    max_c = max(60, int(ws.max_column or 0) + 5)
    for c in range(1, max_c + 1):
        v = ws.cell(row, c).value
        if v is None or str(v).strip() == "":
            empty_run += 1
            if headers and empty_run >= 20:
                break
            continue
        empty_run = 0
        last_non_empty_col = c
        headers.append(str(v).strip())
    return headers


def read_headers_from_template(template_path: str | Path) -> list[str]:
    wb = load_workbook(template_path, data_only=True)
    try:
        ws = wb.active
        return read_header_row(ws, 1)
    finally:
        wb.close()


def _unmerge_rows_below_header(ws, header_rows: int = 1) -> None:
    """Gỡ merge cắt ngang từ hàng dữ liệu trở xuống (openpyxl: MergedCell.value là read-only)."""
    for mrange in list(ws.merged_cells.ranges):
        if mrange.max_row > header_rows:
            ws.unmerge_cells(str(mrange))


def fill_template_from_rows(
    template_path: str | Path,
    output_path: str | Path,
    rows: list[dict[str, Any]],
) -> None:
    """
    Sao chép cấu trúc từ file mẫu và điền dữ liệu từ hàng 2 trở đi.
    rows: mỗi phần tử là dict với key trùng (hoặc gần trùng) tên cột dòng 1.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb.active
    headers = read_header_row(ws, 1)
    if not headers:
        raise ValueError("File mẫu không có dòng tiêu đề (hàng 1).")

    _unmerge_rows_below_header(ws, header_rows=1)
    max_r = max(ws.max_row, 2 + len(rows))
    for r in range(2, max_r + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    col_index = {h: i + 1 for i, h in enumerate(headers)}

    def pick(row: dict[str, Any], header: str) -> Any:
        if header in row:
            return row[header]
        h_norm = _norm_key(header)
        for k, v in row.items():
            if _norm_key(str(k)) == h_norm:
                return v
        return None

    for ridx, row_data in enumerate(rows):
        r = 2 + ridx
        for h in headers:
            val = pick(row_data, h)
            if val is not None:
                ws.cell(r, col_index[h]).value = val

    wb.save(out)


def fill_template_vertical_quiz(
    template_path: str | Path,
    output_path: str | Path,
    blocks: list[VerticalQuizBlock],
) -> None:
    """
    Điền quiz dạng dọc: mỗi câu = 4 hàng (một hàng một đáp án A–D).
    Merge STT, Mức độ, Mục tiêu, Câu hỏi trên 4 hàng của mỗi câu.
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(template_path)
    ws = wb.active
    headers = read_header_row(ws, 1)
    if not is_vertical_quiz_template(headers):
        raise ValueError(
            "File mẫu không đúng khung 7 cột (Mức độ, Mục tiêu, hai «Câu hỏi» hoặc STT+Câu hỏi, "
            "Các đáp án, Kết quả, Giải thích). Bấm «Tạo / mở mẫu mặc định» để lấy file mẫu đúng."
        )
    if len(blocks) != 5:
        raise ValueError("Cần đúng 5 câu cho khung quiz dọc.")

    c_stt, c_md, c_mt, c_q, c_ans, c_res, c_exp = resolve_vertical_column_indexes(headers)
    n_cols = len(headers)

    # Gỡ merge xuống từ hàng 2 (cả merge bắt đầu từ hàng 1 như A1:A50 — nếu không gỡ thì ô là MergedCell, không ghi được).
    _unmerge_rows_below_header(ws, header_rows=1)
    max_clear = max(ws.max_row, 2 + len(blocks) * 4)
    for r in range(2, max_clear + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None

    vcenter = Alignment(vertical="center", wrap_text=True)
    side_thin = Side(style="thin", color="FF000000")
    side_thick = Side(style="medium", color="FF000000")

    for bi, blk in enumerate(blocks):
        r0 = 2 + bi * 4
        # Merge & điền 4 cột meta + câu hỏi
        for c_start, c_end, val in (
            (c_stt, c_stt, blk.stt),
            (c_md, c_md, blk.muc_do),
            (c_mt, c_mt, blk.muc_tieu),
            (c_q, c_q, blk.question_text),
        ):
            ws.merge_cells(
                start_row=r0,
                start_column=c_start,
                end_row=r0 + 3,
                end_column=c_end,
            )
            top = ws.cell(r0, c_start)
            top.value = val
            top.alignment = vcenter
            top.font = Font(name=_QUIZ_FONT_NAME, size=_QUIZ_FONT_SIZE, bold=False)

        labels = ("A", "B", "C", "D")
        for j in range(4):
            r = r0 + j
            text, kq, expl = blk.choices[j]
            prefix = f"{labels[j]}. " if not str(text).lstrip().startswith(f"{labels[j]}.") else ""
            c_ans_cell = ws.cell(r, c_ans)
            c_ans_cell.value = f"{prefix}{text}".strip()
            c_ans_cell.alignment = Alignment(wrap_text=True)
            c_ans_cell.font = Font(name=_QUIZ_FONT_NAME, size=_QUIZ_FONT_SIZE, bold=False)
            _apply_ket_qua_cell(ws.cell(r, c_res), kq)
            c_exp_cell = ws.cell(r, c_exp)
            c_exp_cell.value = clean_quiz_explanation(expl)
            c_exp_cell.alignment = Alignment(wrap_text=True)
            c_exp_cell.font = Font(name=_QUIZ_FONT_NAME, size=_QUIZ_FONT_SIZE, bold=False)

        r_last = r0 + 3
        for c in range(1, n_cols + 1):
            cell = ws.cell(r_last, c)
            if isinstance(cell, MergedCell):
                continue
            b = cell.border
            cell.border = Border(
                left=b.left if b and b.left else side_thin,
                right=b.right if b and b.right else side_thin,
                top=b.top if b and b.top else side_thin,
                bottom=side_thick,
            )

    # Hàng tiêu đề: font + wrap + autofit để cột Giải thích / Câu hỏi đủ rộng
    for c in range(1, n_cols + 1):
        hcell = ws.cell(1, c)
        if not isinstance(hcell, MergedCell):
            hcell.font = Font(name=_QUIZ_FONT_NAME, size=_QUIZ_FONT_SIZE, bold=True)
            hcell.alignment = Alignment(wrap_text=True, vertical="center")

    autofit_quiz_columns_and_rows(
        ws,
        header_row=1,
        vertical_blocks=blocks,
        c_q=c_q,
        c_ans=c_ans,
        c_res=c_res,
        c_exp=c_exp,
    )

    wb.save(out)
